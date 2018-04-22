# -*- coding: utf-8 -*-
"""User models."""
import datetime as dt

from database import ModelBase, dbsession, reference_col
from sqlalchemy import column, Column, ForeignKey, Integer, String, Text, Boolean, DateTime, Date, Table, MetaData, select
from sqlalchemy.orm import relationship, backref
from sqlalchemy.sql import text
from template import render_string
from users.models import ModuleRole, User
from pages.models import Page
from openpyxl import load_workbook
import json
import math

class Tracker(ModelBase):
    __tablename__ = 'trackers'
    id = Column(Integer, primary_key=True)
    title = Column(String(200))
    slug = Column(String(100),unique=True)
    module = Column(String(100),default='pages')
    list_fields = Column(Text)
    search_fields = Column(Text)
    pagelimit = Column(Integer,default=10)

    def __repr__(self):
        return self.title

    def list_fields_list(self):
        rfields = []
        if self.list_fields:
            pfields = self.list_fields.split(',')
            for pfield in pfields:
                rfield = dbsession.query(TrackerField).filter_by(tracker=self,name=pfield.strip()).first()
                if rfield:
                    rfields.append(rfield)
        return rfields

    def display_fields_list(self,record):
        curstatus = self.status(record)
        if curstatus:
            return self.fields_from_list(curstatus.display_fields)
        return []

    def fields_from_list(self,field_list=None):
        rfields = []
        if field_list:
            pfields = field_list.split(',')
            rfields = []
            for pfield in pfields:
                rfield = dbsession.query(TrackerField).filter_by(tracker=self,name=pfield.strip()).first()
                if rfield:
                    rfields.append(rfield)
        return rfields

    def data_table(self):
        return "trak_" + self.slug + "_data"

    def update_table(self):
        return "trak_" + self.slug + "_updates"

    @property
    def pages(self):
        return dbsession.query(Page).filter_by(module=self.module).all()

    def field(self, name):
        field = dbsession.query(TrackerField).filter_by(tracker=self,name=name.strip()).first()
        return field

    def updatedb(self):
        query = """
            do $$
            begin
            if (select not exists(select  1 from information_schema.sequences where sequence_schema = 'public' and sequence_name = '""" + self.data_table() + """_id_seq'))
            then
                create sequence public.""" + self.data_table() + """_id_seq
                increment 1
                minvalue 1
                maxvalue 9223372036854775807
                start 1
                cache 1;
            end if;
            if (select not exists(select  1 from information_schema.tables where table_schema = 'public' and table_name = '""" + self.data_table() + """'))
            then
                create table public.""" + self.data_table() + """(
                    id integer not null default nextval('""" + self.data_table() + """_id_seq'::regclass),
                    record_status character varying(50) ,
                    batch_no integer
                );
            end if;
            if (select not exists(select  1 from information_schema.sequences where sequence_schema = 'public' and sequence_name = '""" + self.update_table() + """_id_seq'))
            then
                create sequence public.""" + self.update_table() + """_id_seq
                increment 1
                minvalue 1
                maxvalue 9223372036854775807
                start 1
                cache 1;
            end if;
            if (select not exists(select  1 from information_schema.tables where table_schema = 'public' and table_name = '""" + self.update_table() + """'))
            then
                create table public.""" + self.update_table() + """(
                    id integer not null default nextval('""" + self.update_table() + """_id_seq'::regclass),
                    record_id integer,
                    user_id integer,
                    record_status character varying(50),
                    update_date timestamp,
                    description text
                );
            end if;
            end$$;
        """
        dbsession.execute(query)
        try:
            dbsession.commit()
        except Exception as inst:
            dbsession.rollback()
        for field in self.fields:
            field.updatedb()

    def saverecord(self, record, request=None):
        curuser = None
        data = None
        if request:
            if 'user_id' in request['session']:
                curuser = dbsession.query(User).filter(User.id==request['session']['user_id']).first()
        if 'id' in record:
            query = """update """ + self.data_table() + """ set """ + ",".join([ formfield + "=" + self.field(formfield).sqlvalue(record[formfield]) for formfield in record.keys()  ]) + """ where id=""" + str(record['id']) + " returning *"
        else:
            query = """
                insert into """ + self.data_table() + """ ( """ + ",".join(record.keys()) + """) values 
                (""" + ",".join([ self.field(formfield).sqlvalue(record[formfield]) for formfield in record.keys()  ]) + """) returning *
            """
        try:
            data = dbsession.execute(query).fetchone()
            dbsession.commit()
        except Exception as inst:
            dbsession.rollback()
        return data

    def addrecord(self, form, request):
        curuser = None
        data = None
        if 'user_id' in request['session']:
            curuser = dbsession.query(User).filter(User.id==request['session']['user_id']).first()
        if 'transition_id' in form:
            transition = dbsession.query(TrackerTransition).get(form['transition_id'][0])
            if transition:
                if transition.next_status:
                    form['record_status'] = [transition.next_status.name,]
                for field in transition.edit_fields_list():
                    if field.default and field.name in form and form[field.name][0]=='systemdefault':
                        output=None
                        ldict = locals()
                        exec(field.default,globals(),ldict)
                        output=ldict['output']
                        form[field.name][0]=output
                del(form['transition_id'])
        fieldnames = list(form.keys())
        query = """
            insert into """ + self.data_table() + """ ( """ + ",".join(fieldnames) + """) values 
            (""" + ",".join([ self.field(formfield).sqlvalue(form[formfield][0]) for formfield in fieldnames  ]) + """) returning *
        """
        try:
            data = dbsession.execute(query).fetchone()
            dbsession.commit()
        except Exception as inst:
            dbsession.rollback()
        desc = ''
        if curuser:
            desc = 'Updated by ' + curuser.name
        else:
            desc = 'Updated by anonymous'

        query = """
            insert into """ + self.update_table() + """ (record_id,user_id,record_status,update_date,description) values 
            ( """ + str(data['id']) + "," + (str(curuser.id) + "," if curuser else 'null,') + "'" + data['record_status'] + "',now(),'" + desc + "') "
        try:
            dbsession.execute(query)
            dbsession.commit()
        except Exception as inst:
            dbsession.rollback()

        return data

    def editrecord(self, form, request, id=None):
        curuser = None
        if 'user_id' in request['session']:
            curuser = dbsession.query(User).filter(User.id==request['session']['user_id']).first()
        if 'transition_id' in form:
            transition = dbsession.query(TrackerTransition).get(form['transition_id'][0])
            if transition and transition.next_status:
                form['record_status'] = [transition.next_status.name,]
            del(form['transition_id'])
        oldrecord = None
        if id:
            oldrecord = self.records(id,curuser=curuser,request=request)
        elif 'id' in form:
            oldrecord = self.records(form['id'][0],curuser=curuser,request=request)
            del(form['id'])
        fieldnames = list(form.keys())
        data = None
        if oldrecord:
            query = """update """ + self.data_table() + """ set """ + ",".join([ formfield + "=" + self.field(formfield).sqlvalue(form[formfield][0]) for formfield in fieldnames  ]) + """ where id=""" + str(oldrecord['id']) + " returning *"
        try:
            data = dbsession.execute(query).fetchone()
            dbsession.commit()
        except Exception as inst:
            dbsession.rollback()
        desc = ''
        if curuser:
            desc = 'Updated by ' + curuser.name
        else:
            desc = 'Updated by anonymous'
        if data:
            query = """
                insert into """ + self.update_table() + """ (record_id,user_id,record_status,update_date,description) values 
                ( """ + str(data['id']) + "," + (str(curuser.id) + "," if curuser else 'null,') + "'" + data['record_status'] + "',now(),'" + desc + "') "
            try:
                dbsession.execute(query)
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()

        return data

    def pagelinks(self,curuser=None,request=None,cleared=False):
        recordamount = dbsession.execute("select count(*) as num from " + self.data_table() + self.queryrules(curuser=curuser, request=request)).first()['num']
        pageamount = int(math.ceil(recordamount/(self.pagelimit if self.pagelimit else 10)))
        links = []
        curoffset = 0
        if request:
            curoffset = int(request.args['plo'][0]) if 'plo' in request.args else 0
        curindex = 0
        for x in range(0,pageamount):
            nextoffset=((x+1)*(self.pagelimit if self.pagelimit else 10) if x+1<pageamount else None)
            prevoffset=((x-1)*(self.pagelimit if self.pagelimit else 10) if x else None)
            offset=x*(self.pagelimit if self.pagelimit else 10)
            if 'plq' in request.args:
                url=request.app.url_for('trackers.viewlist',plq=request.args['plq'][0],slug=self.slug,plo=offset, pll=(self.pagelimit if self.pagelimit else 10))
                prevlink = (request.app.url_for('trackers.viewlist',plq=request.args['plq'][0],slug=self.slug,plo=prevoffset, pll=(self.pagelimit if self.pagelimit else 10))) if prevoffset else None
                nextlink = (request.app.url_for('trackers.viewlist',plq=request.args['plq'][0],slug=self.slug,plo=nextoffset, pll=(self.pagelimit if self.pagelimit else 10))) if nextoffset else None
            else:
                url=request.app.url_for('trackers.viewlist',slug=self.slug,plo=offset, pll=(self.pagelimit if self.pagelimit else 10))
                prevlink = (request.app.url_for('trackers.viewlist',slug=self.slug,plo=prevoffset, pll=(self.pagelimit if self.pagelimit else 10))) if prevoffset else None
                nextlink = (request.app.url_for('trackers.viewlist',slug=self.slug,plo=nextoffset, pll=(self.pagelimit if self.pagelimit else 10))) if nextoffset else None
            thislink = { 'url':url,'nextlink':nextlink,'prevlink':prevlink }
            if curoffset==x*(self.pagelimit if self.pagelimit else 10):
                curindex = x
            links.append(thislink)
        return curindex,links

    def queryrules(self,curuser=None,request=None,cleared=False):
        rules = ' where 1=1 '
        if 'plq' in request.args: # page list query exists
            if self.search_fields and len(request.args['plq'][0]):
                sfields = self.fields_from_list(self.search_fields)
                sqs = []
                for sfield in sfields:
                    sqs.append(sfield.queryvalue(request.args['plq'][0]))
                rules += ' and (' + ' or '.join(sqs) + ')'

        if not cleared:
            rrules = self.rolesrule(curuser,request)
            if rrules:
                rules += ' and (' + rrules + ') '
        return rules

    def records(self,id=None,curuser=None,request=None,cleared=False,offset=None,limit=None):
        results = []
        if id:
            if cleared:
                sqltext = text("select * from " + self.data_table() +  " where id=:id")
            else:
                sqltext = text("select * from " + self.data_table() +  " where id=:id and (" + self.rolesrule(curuser,request) + ")")
            sqltext = sqltext.bindparams(id=id)
        else:
            limitstr = ''
            offsetstr = ''
            if limit:
                limitstr = ' limit ' + str(limit)
            if offset:
                offsetstr = ' offset ' + str(offset)
            sqltext = text("select * from " + self.data_table() + self.queryrules(curuser=curuser,request=request,cleared=cleared) + limitstr + offsetstr)
        try:
            results = dbsession.execute(sqltext)
        except Exception as inst:
            dbsession.rollback()
        if id:
            drows = None
            for row in results:
                drows = row
            results = drows
        return results

    def rolesrule(self,curuser,request):
        if self.userroles(curuser):
            rolesrule = '1=1'
        else:
            queryroles = dbsession.query(TrackerRole).filter_by(tracker=self,role_type='query').all()
            rolesrule = '1=0'
            rulestext = [ render_string(request,role.compare) for role in queryroles ]
            if len(rulestext):
                trolesrule = ' or '.join(rulestext)
                if len(trolesrule):
                    rolesrule = trolesrule
        return rolesrule

    def status(self,record):
        if record and 'record_status' in record and record['record_status']:
            status = dbsession.query(TrackerStatus).filter_by(tracker=self,name=record['record_status']).first()
            if status:
                return status
        return None

    def userroles(self,curuser,record=None,request=None):
        croles = []
        for role in self.roles:
            if role.role_type=='module':
                usermoduleroles = dbsession.query(ModuleRole).filter(ModuleRole.module==self.module,ModuleRole.user==curuser,ModuleRole.role==role.name).all()
                if len(usermoduleroles):
                    croles.append(role)
            elif record:
                rolesrule = render_string(request,role.compare)
                sqltext = "select id from " + self.data_table() + " where id=" + str(record['id']) + " and " + rolesrule
                results = dbsession.execute(sqltext)
                if results:
                    for row in results:
                        croles.append(role)
        return croles

    def activetransitions(self,record,curuser,request):
        status = self.status(record)
        roles = self.userroles(curuser,record,request=request)
        atransitions = []
        if status:
            transitions = dbsession.query(TrackerTransition).filter(TrackerTransition.prev_status==status,TrackerTransition.tracker==self).all()
            if len(transitions) and len(roles):
                for t in transitions:
                    for r in roles:
                        if r in t.roles:
                            atransitions.append(t)
        return atransitions


class TrackerField(ModelBase):
    __tablename__ = 'tracker_fields'
    id = Column(Integer, primary_key=True)
    name = Column(String(50))
    label = Column(String(50))
    field_type = Column(String(20))
    obj_table = Column(String(50))
    obj_field = Column(String(100))
    default = Column(Text())

    tracker_id = reference_col('trackers')
    tracker = relationship('Tracker',backref='fields')

    def __repr__(self):
        return self.label

    def obj_fields(self):
        if self.obj_field:
            return self.obj_field.split(',')

    def main_obj_field(self):
        if self.obj_field:
            return self.obj_fields()[0]

    def disp_value(self, value):
        if value:
            if self.field_type=='object':
                sqlq = "select " + self.main_obj_field() + " from " + self.obj_table + " where id=" + str(value)
                result = dbsession.execute(sqlq)
                for r in result:
                    return r[0]
            elif self.field_type=='user':
                sqlq = "select name from users where id=" + str(value)
                result = dbsession.execute(sqlq)
                for r in result:
                    return r[0]
            
        return value

    def queryvalue(self, value):
        if self.field_type in ['string','text']:
            return self.name + " ilike '%" + str(value) + "%'"
        elif self.field_type in ['integer','number','object','user']:
            return self.name + "=" + str(value)
        elif self.field_type in ['date','datetime']:
            return self.name + "='" + str(value) + "'"
        return "1=1"

    def sqlvalue(self, value):
        if value:
            if self.field_type in ['string','text','date','datetime']:
                return "'" + str(value).replace("'","''") + "'"
            elif self.field_type in ['integer','number','object','user']:
                return str(value)
            elif self.field_type=='boolean':
                if value:
                    return str(1)
                else:
                    return str(0)
        else:
            return "null"
        return ''

    def dbcolumn(self):
        return column(self.name)

    def db_field_type(self):
        if self.field_type=='string':
            return 'character varying(200)'
        elif self.field_type=='text':
            return 'text'
        elif self.field_type=='integer':
            return 'integer'
        elif self.field_type=='object':
            return 'integer'
        elif self.field_type=='user':
            return 'integer'
        elif self.field_type=='number':
            return 'double precision'
        elif self.field_type=='date':
            return 'date'
        elif self.field_type=='datetime':
            return 'timestamp'
        elif self.field_type=='boolean':
            return 'boolean'

    def updatedb(self):
        query = """
            do $$
            begin 
                IF not EXISTS (SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_schema='public' and table_name='""" + self.tracker.data_table() + """' and column_name='""" + self.name + """') THEN
                        alter table public.""" + self.tracker.data_table() + """ add column """ + self.name + " " + self.db_field_type() + """ null ;
                end if;
            end$$;
        """
        dbsession.execute(query)
        dbsession.commit()

class TrackerRole(ModelBase):
    __tablename__ = 'tracker_roles'
    id = Column(Integer, primary_key=True)
    name = Column(String(50))
    role_type = Column(String(10))
    compare = Column(Text,nullable=True)

    tracker_id = reference_col('trackers')
    tracker = relationship('Tracker',backref='roles')

    def __repr__(self):
        return self.name

class TrackerDataUpdate(ModelBase):
    __tablename__ = 'tracker_data_update'
    id = Column(Integer, primary_key=True)
    status = Column(String(50))
    created_date = Column(DateTime)
    filename = Column(String(200))
    data_params = Column(Text,nullable=True)

    tracker_id = reference_col('trackers')
    tracker = relationship('Tracker',backref='dataupdates')

    def __repr__(self):
        return self.tracker.title + ' ' + str(self.created_date)

    async def run(self):
        wb = load_workbook(filename = self.filename)
        ws = wb.active
        datas = json.loads(self.data_params)
        fields = []
        optype = 'insert'
        searchfield = []
        for field in self.tracker.fields:
            if field.name != 'id' and datas[field.name + '_column'][0]!='ignore':
                fields.append(field)
            if field.name + '_update' in datas:
                optype = 'update'
                searchfield.append(field)
        rows = []
        headerend = 1
        if optype == 'insert':
            query = 'insert into ' + self.tracker.data_table() + ' (' + ','.join([ f.name for f in fields ]) + ') values '
            for i,row in enumerate(ws.rows):
                if i>headerend:
                    cellrows = [ws[datas[f.name + '_column'][0] + str(i+1)] for f in fields]
                    cellpos = [ datas[f.name + '_column'][0] + str(i+1) for f in fields ]
                    drowdata = [ f.value for f in cellrows ]
                    sqlrowdata = [ f.sqlvalue(drowdata[dd]) for dd,f in enumerate(fields) ]
                    drow = '('
                    drow = drow + ','.join(sqlrowdata)
                    drow = drow + ')'
                    rows.append(drow)
            query = query + ','.join(rows)
            try:
                dbsession.execute(query)
                dbsession.commit()
                self.status = 'Uploaded ' + str(len(rows)) + ' rows'
                dbsession.add(self)
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
        else:
            allresults = []
            for i,row in enumerate(ws.rows):
                if i>headerend:
                    fieldupdates = []
                    for f in searchfield:
                        cellrow = ws[datas[f.name + '_column'][0] + str(i+1)]
                        fieldupdates.append( f.name + '=' + f.sqlvalue(cellrow.value) )
                    queryrow = ' where ' + 'and'.join(fieldupdates)

                    sqltext = text("select id from " + self.tracker.data_table() + queryrow)
                    result = dbsession.execute(sqltext)
                    gotdata = False
                    for r in result:
                        gotdata = True
                    
                    if gotdata:
                        query = 'update ' + self.tracker.data_table() + ' set ' 
                        fieldinfos = []
                        for f in fields:
                            if not f in searchfield:
                                cellrow = ws[datas[f.name + '_column'][0] + str(i+1)]
                                fieldinfos.append( f.name + '=' + f.sqlvalue(cellrow.value) )
                        query += ','.join(fieldinfos) + queryrow + ' returning *'
                        result = dbsession.execute(query)
                    else:
                        cellrow = ws[datas[f.name + '_column'][0] + str(i+1)]
                        query = 'insert into ' + self.tracker.data_table() + ' (' + ','.join([ f.name for f in fields ]) + ') values (' + ','.join([ f.sqlvalue(cellrow.value) for f in fields ]) + ')'
                        query += ' returning *'
                        result = dbsession.execute(query)
                    try:
                        dbsession.commit()
                        allresults.append('Uploaded ' + str(result.first()))
                    except Exception as inst:
                        dbsession.rollback()
            try:
                self.status = 'Updated ' + str(len(allresults)) + ' rows'
                dbsession.add(self)
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
        return True

class TrackerStatus(ModelBase):
    __tablename__ = 'tracker_statuses'
    id = Column(Integer, primary_key=True)
    name = Column(String(50))
    display_fields = Column(Text)

    tracker_id = reference_col('trackers')
    tracker = relationship('Tracker',backref='statuses')

    def __repr__(self):
        return self.name

transition_roles = Table('transition_roles',ModelBase.metadata,
        Column('transition_id',Integer,ForeignKey('tracker_transitions.id')),
        Column('role_id',Integer,ForeignKey('tracker_roles.id'))
        )

class TrackerTransition(ModelBase):
    __tablename__ = 'tracker_transitions'
    id = Column(Integer, primary_key=True)
    name = Column(String(50))
    display_fields = Column(Text)
    edit_fields = Column(Text)
    postpage = Column(Text)

    roles = relationship('TrackerRole',secondary=transition_roles,backref=backref('transitions',lazy='dynamic'))

    tracker_id = reference_col('trackers')
    tracker = relationship('Tracker',backref='transitions')

    prev_status_id = reference_col('tracker_statuses',nullable=True)
    prev_status = relationship('TrackerStatus',backref='prev_transitions',foreign_keys=[prev_status_id])

    next_status_id = reference_col('tracker_statuses',nullable=True)
    next_status = relationship('TrackerStatus',backref='next_transitions',foreign_keys=[next_status_id])

    def __repr__(self):
        return self.name

    def edit_fields_list(self):
        pfields = self.edit_fields.split(',') if self.edit_fields else []
        rfields = []
        for pfield in pfields:
            rfield = dbsession.query(TrackerField).filter_by(tracker=self.tracker,name=pfield.strip()).first()
            if rfield:
                rfields.append(rfield)
        return rfields

    def display_fields_list(self):
        pfields = self.display_fields.split(',')
        rfields = []
        for pfield in pfields:
            rfield = dbsession.query(TrackerField).filter_by(tracker=self.tracker,name=pfield.strip()).first()
            if rfield:
                rfields.append(rfield)
        return rfields
