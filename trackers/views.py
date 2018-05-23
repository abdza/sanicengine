# -*- coding: utf-8 -*-
from sanic import Blueprint
from sanic.response import html, redirect, json as jsonresponse, file_stream, stream, raw
from .models import Tracker, TrackerField, TrackerRole, TrackerStatus, TrackerTransition, TrackerDataUpdate
from users.models import User
from pages.models import Page
from .forms import TrackerForm, TrackerFieldForm, TrackerRoleForm, TrackerStatusForm, TrackerTransitionForm
from database import dbsession
from template import render
from decorators import authorized
from sqlalchemy_paginator import Paginator
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
import os
import datetime
import json
import re
import asyncio
import io
from openpyxl import load_workbook, Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook


bp = Blueprint('trackers')

uploadfolder = 'upload'

def slugify(slug):
    slug = re.sub('[^0-9a-zA-Z]+', '_', slug.lower())
    return slug

@bp.route('/trackers/runupdate/<module>',methods=['GET'])
@bp.route('/trackers/runupdate/<module>/<slug>',methods=['GET'])
async def runupdate(request,module,slug=None):
    if slug==None:
        slug=module
        module='portal'
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    updates = dbsession.query(TrackerDataUpdate).filter_by(status='new').all()
    for update in updates:
        update.status = 'in queue'
        dbsession.add(update)
    try:
        dbsession.commit()
    except Exception as inst:
        dbsession.rollback()

    try:
        await asyncio.wait([ update.run() for update in updates])
    except:
        print("Nothing to wait for")
    return redirect('/trackers/view/' + str(tracker.id) + '#dataupdates')

@bp.route('/trackers/deleteupdate/<update_id>',methods=['POST'])
@authorized(object_type='dataupdate')
async def deleteupdate(request,update_id=None):
    if update_id:
        update = dbsession.query(TrackerDataUpdate).get(int(update_id))
        trackerid = update.tracker.id
        if(update):
            if update.filename and os.path.exists(update.filename):
                os.remove(update.filename)
            if os.path.exists(os.path.join(uploadfolder,update.tracker.slug,'dataupdate',str(update.id))):
                os.rmdir(os.path.join(uploadfolder,update.tracker.slug,'dataupdate',str(update.id)))
            dbsession.delete(update)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
        return redirect('/trackers/view/' + str(trackerid) + '#dataupdates')
    else:
        request['session']['flashmessage']='Need to specify id of update to delete'
        return redirect(request.app.url_for('trackers.index'))

@bp.route('/trackers/dataupdate/<module>',methods=['GET','POST'])
@bp.route('/trackers/dataupdate/<module>/<slug>',methods=['GET','POST'])
@authorized(object_type='dataupdate')
async def data_update(request,module,slug=None):
    if slug==None:
        slug=module
        module='portal'
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    columns = []
    dataupdate = None
    if request.method=='POST':
        if request.form.get('dataupdate_id'):
            dataupdate = dbsession.query(TrackerDataUpdate).get(str(request.form.get('dataupdate_id')))
            dataupdate.data_params=json.dumps(request.form)
            dbsession.add(dataupdate)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            return redirect('/trackers/view/' + str(tracker.id) + '#dataupdates')
        if request.files.get('excelfile'):
            dataupdate = TrackerDataUpdate(status='new',tracker=tracker,created_date=datetime.datetime.now())
            dbsession.add(dataupdate)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            dfile = request.files.get('excelfile')
            ext = dfile.type.split('/')[1]
            if not os.path.exists(os.path.join(uploadfolder,tracker.slug,'dataupdate',str(dataupdate.id))):
                os.makedirs(os.path.join(uploadfolder,tracker.slug,'dataupdate',str(dataupdate.id)))
            dst = os.path.join(uploadfolder,tracker.slug,'dataupdate',str(dataupdate.id),dfile.name)
            try:
                # extract starting byte from Content-Range header string
                range_str = request.headers['Content-Range']
                start_bytes = int(range_str.split(' ')[1].split('-')[0])
                with open(dst, 'ab') as f:
                    f.seek(start_bytes)
                    f.write(dfile.body)
            except KeyError:
                with open(dst, 'wb') as f:
                    f.write(dfile.body)
            dataupdate.filename = dst
            dbsession.add(dataupdate)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            wb = load_workbook(filename = dst)
            ws = wb.active
            columns.append({'field_val':'ignore','field_name':'Ignore'})
            columns.append({'field_val':'custom','field_name':'Custom'})
            for row in ws.iter_rows(max_row=1):
                for cell in row:
                    if cell.value.lower() != 'id':
                        columns.append({'field_val':cell.column,'field_name':cell.value})
    return html(render(request,'/trackers/data_update.html',tracker=tracker,columns=columns,dataupdate=dataupdate))

@bp.route('/trackers/create_from_excel/<module>',methods=['POST','GET'])
@bp.route('/trackers/create_from_excel/<module>/<slug>',methods=['POST','GET'])
@authorized(require_admin=True)
async def create_from_excel(request,module,slug=None):
    if slug==None:
        slug=module
        module='portal'
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    fields = []
    field_types = []
    if request.method=='POST':
        if request.form.get('field_name'):
            field_names = request.form['field_name']
            field_label = request.form['field_label']
            field_type = request.form['field_type']
            for i,name in enumerate(field_names):
                tfield = TrackerField(tracker=tracker, name=name, label=field_label[i], field_type=field_type[i])
                dbsession.add(tfield)
            tracker.list_fields = ','.join(field_names)
            dbsession.add(tracker)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            return redirect('/trackers/view/' + str(tracker.id) + '#fields')
        if request.files.get('excelfile'):
            dfile = request.files.get('excelfile')
            ext = dfile.type.split('/')[1]
            dst = os.path.join(uploadfolder,dfile.name)
            if not os.path.exists(uploadfolder):
                os.makedirs(uploadfolder)
            try:
                # extract starting byte from Content-Range header string
                range_str = request.headers['Content-Range']
                start_bytes = int(range_str.split(' ')[1].split('-')[0])
                with open(dst, 'ab') as f:
                    f.seek(start_bytes)
                    f.write(dfile.body)
            except KeyError:
                with open(dst, 'wb') as f:
                    f.write(dfile.body)
            wb = load_workbook(filename = dst)
            ws = wb.active
            fieldtitles = []
            fieldtypes = []
            for row in ws.iter_rows(max_row=1):
                for cell in row:
                    fieldtitles.append(cell.value)
            for row in ws.iter_rows(max_row=1,row_offset=1):
                for cell in row:
                    fieldtypes.append(cell.data_type)
            for i,title in enumerate(fieldtitles):
                if slugify(title)!='id':
                    fields.append({'field_name':slugify(title),'field_type':fieldtypes[i]})
            os.remove(dst)
        field_types = [('string','String'),('text','Text'),('integer','Integer'),('number','Number'),('date','Date'),('datetime','Date Time'),('boolean','Boolean'),('object','Object')]
    return html(render(request,'/trackers/create_from_excel.html',tracker=tracker,fields=fields,field_types=field_types))

@bp.route('/trackers/<module>/<slug>/addstatus',methods=['POST','GET'])
@bp.route('/trackers/<module>/<slug>/editstatus/',methods=['POST','GET'],name='editstatus')
@bp.route('/trackers/<module>/<slug>/editstatus/<id>',methods=['POST','GET'],name='editstatus')
@authorized(require_admin=True)
async def statusform(request,module,slug,id=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    title = 'Create Tracker Status'
    form = TrackerStatusForm(request.form)
    trackerstatus = None
    if id:
        trackerstatus = dbsession.query(TrackerStatus).get(int(id))
    if trackerstatus:
        title = 'Edit Tracker Status'
    tokeninput = {
            'display_fields': {
                'url': request.app.url_for('trackers.trackerfieldsjson',module=module,slug=slug),
                },
            }
    if request.method=='POST':
        if form.validate():
            if not trackerstatus:
                trackerstatus=TrackerStatus()
            form.populate_obj(trackerstatus)
            trackerstatus.tracker = tracker
            if form['display_fields'].data:
                trackerstatus.display_fields = ','.join([ dbsession.query(TrackerField).get(int(fieldid)).name for fieldid in form['display_fields'].data.split(',') ])
            dbsession.add(trackerstatus)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            return redirect('/trackers/view/' + str(trackerstatus.tracker.id) + '#statuses')
    else:
        if id:
            trackerstatus = dbsession.query(TrackerStatus).get(int(id))
        if trackerstatus:
            form = TrackerStatusForm(obj=trackerstatus)
            title = 'Edit Tracker Status'
            tokeninput = {
                    'display_fields': {
                        'url': request.app.url_for('trackers.trackerfieldsjson',module=trackerstatus.tracker.module,slug=trackerstatus.tracker.slug),
                        'prePopulate':[ {'id':field.id,'name':field.name} for field in trackerstatus.tracker.fields_from_list(trackerstatus.display_fields) ]
                        },
                    }
    return html(render(request,'generic/form.html',title=title,form=form,enctype='multipart/form-data',tokeninput=tokeninput))

@bp.route('/trackers/deletestatus/<status_id>',methods=['POST'])
@authorized(require_admin=True)
async def deletestatus(request,status_id=None):
    if status_id:
        trackerstatus = dbsession.query(TrackerStatus).get(int(status_id))
        trackerid = trackerstatus.tracker.id
        if(trackerstatus):
            dbsession.delete(trackerstatus)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
        return redirect('/trackers/view/' + str(trackerid) + '#statuses')
    else:
        request['session']['flashmessage']='Need to specify id of status to delete'
        redirect(request.app.url_for('trackers.index'))

@bp.route('/trackers/<module>/<slug>/roles/json')
async def rolesjson(request,module,slug=None):
    if slug==None:
        slug=module
        module='portal'
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    trackerroles = dbsession.query(TrackerRole).filter(TrackerRole.tracker==tracker,TrackerRole.name.ilike('%' + request.args['q'][0] + '%')).all() 
    return jsonresponse([ {'id':role.id,'name':role.name} for role in trackerroles ])

@bp.route('/trackers/<module>/<slug>/fields/json')
async def trackerfieldsjson(request,module,slug=None):
    if slug==None:
        slug=module
        module='portal'
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    trackerfields = dbsession.query(TrackerField).filter(TrackerField.tracker==tracker,TrackerField.name.ilike('%' + request.args['q'][0] + '%')).all() 
    return jsonresponse([ {'id':field.id,'name':field.name} for field in trackerfields ])

@bp.route('/trackers/<module>/<slug>/addtransition',methods=['POST','GET'])
@bp.route('/trackers/<module>/<slug>/edittransition/',methods=['POST','GET'],name='edittransition')
@bp.route('/trackers/<module>/<slug>/edittransition/<id>',methods=['POST','GET'],name='edittransition')
@authorized(require_admin=True)
async def transitionform(request,module,slug=None,id=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    title = 'Create Tracker Transition'
    form = TrackerTransitionForm(request.form)
    dstatuses = [('','None'),] + [(str(g.id),g.name) for g in dbsession.query(TrackerStatus).filter(TrackerStatus.tracker==tracker).all()]
    form.prev_status_id.choices = dstatuses
    form.next_status_id.choices = dstatuses
    trackertransition = None
    tokeninput = {
            'display_fields': {
                'url': request.app.url_for('trackers.trackerfieldsjson',module=module,slug=slug),
                },
            'edit_fields': {
                'url': request.app.url_for('trackers.trackerfieldsjson',module=module,slug=slug),
                },
            'roles': {
                'url': request.app.url_for('trackers.rolesjson',module=module,slug=slug),
                },
            }
    if id:
        trackertransition = dbsession.query(TrackerTransition).get(int(id))
    if trackertransition:
        title = 'Edit Tracker Transition'
    if request.method=='POST':
        if form.validate():
            if not trackertransition:
                trackertransition=TrackerTransition()
            if form.roles.data:
                role_ids = form.roles.data.split(',')
                roles = [ dbsession.query(TrackerRole).get(role_id) for role_id in role_ids ]
                trackertransition.roles = roles
            else:
                trackertransition.roles = []
            del(form['roles'])
            display_fields = []
            edit_fields = []
            if form['display_fields'].data:
                display_fields = ','.join([ ddat.name for ddat in dbsession.execute("select name from tracker_fields where id in (" + form['display_fields'].data + ")") ])
                del(form['display_fields'])
            if form['edit_fields'].data:
                edit_fields = ','.join([ ddat.name for ddat in dbsession.execute("select name from tracker_fields where id in (" + form['edit_fields'].data + ")") ])
                del(form['edit_fields'])
            form.populate_obj(trackertransition)
            trackertransition.display_fields = display_fields
            trackertransition.edit_fields = edit_fields
            if form.prev_status_id.data=='':
                trackertransition.prev_status = None
                trackertransition.prev_status_id = None
                del(form['prev_status_id'])
            if form.next_status_id.data=='':
                trackertransition.next_status = None
                trackertransition.next_status_id = None
                del(form['next_status_id'])
            form.populate_obj(trackertransition)
            trackertransition.tracker = tracker
            dbsession.add(trackertransition)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            return redirect('/trackers/view/' + str(trackertransition.tracker.id) + '#transitions')
    else:
        if id:
            trackertransition = dbsession.query(TrackerTransition).get(int(id))
        if trackertransition:
            form = TrackerTransitionForm(obj=trackertransition)
            form.prev_status_id.choices = dstatuses
            form.next_status_id.choices = dstatuses
            title = 'Edit Tracker Transition'
            tokeninput['display_fields']['prePopulate'] = [ {'id':field.id,'name':field.name} for field in trackertransition.tracker.fields_from_list(trackertransition.display_fields) ]
            tokeninput['edit_fields']['prePopulate'] = [ {'id':field.id,'name':field.name} for field in trackertransition.tracker.fields_from_list(trackertransition.edit_fields) ]
            tokeninput['roles']['prePopulate'] = [ {'id':field.id,'name':field.name} for field in trackertransition.roles ]
    return html(render(request,'generic/form.html',title=title,form=form,tracker=tracker,enctype='multipart/form-data',tokeninput=tokeninput))

@bp.route('/trackers/deletetransition/<transition_id>',methods=['POST'])
@authorized(require_admin=True)
async def deletetransition(request,transition_id=None):
    if transition_id:
        trackertransition = dbsession.query(TrackerTransition).get(int(transition_id))
        trackerid=trackertransition.tracker.id
        if(trackertransition):
            dbsession.delete(trackertransition)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
        return redirect('/trackers/view/' + str(trackerid) + '#transitions')
    else:
        request['session']['flashmessage']='You need to specify id of transition to delete'
        return redirect(request.app.url_for('trackers.index'))

@bp.route('/trackers/<module>/<slug>/addrole',methods=['POST','GET'])
@bp.route('/trackers/<module>/<slug>/editrole/',methods=['POST','GET'],name='editrole')
@bp.route('/trackers/<module>/<slug>/editrole/<id>',methods=['POST','GET'],name='editrole')
@authorized(require_admin=True)
async def roleform(request,module,slug=None,id=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    title = 'Create Tracker Role'
    form = TrackerRoleForm(request.form)
    trackerrole = None
    if id:
        trackerrole = dbsession.query(TrackerRole).get(int(id))
    if trackerrole:
        title = 'Edit Tracker Role'
    if request.method=='POST':
        if form.validate():
            if not trackerrole:
                trackerrole=TrackerRole()
            form.populate_obj(trackerrole)
            trackerrole.tracker = tracker
            dbsession.add(trackerrole)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            return redirect('/trackers/view/' + str(trackerrole.tracker.id) + '#roles')
    else:
        if id:
            trackerrole = dbsession.query(TrackerRole).get(int(id))
        if trackerrole:
            form = TrackerRoleForm(obj=trackerrole)
            title = 'Edit Tracker Role'
    return html(render(request,'generic/form.html',title=title,form=form,enctype='multipart/form-data'))

@bp.route('/trackers/deleterole/<role_id>',methods=['POST'])
@authorized(require_admin=True)
async def deleterole(request,role_id=None):
    if role_id:
        trackerrole = dbsession.query(TrackerRole).get(int(role_id))
        trackerid = trackerrole.tracker.id
        if(trackerrole):
            dbsession.delete(trackerrole)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
        return redirect('/trackers/view/' + str(trackerid) + '#roles')
    else:
        request['session']['flashmessage']='You need to specify id of role to delete'
        return redirect(request.app.url_for('trackers.index'))

@bp.route('/trackers/<module>/<slug>/addfield',methods=['POST','GET'])
@bp.route('/trackers/<module>/<slug>/editfield/',methods=['POST','GET'],name='editfield')
@bp.route('/trackers/<module>/<slug>/editfield/<id>',methods=['POST','GET'],name='editfield')
@authorized(require_admin=True)
async def fieldform(request,module,slug=None,id=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    title = 'Create Tracker Field'
    form = TrackerFieldForm(request.form)
    trackerfield = None
    if id:
        trackerfield = dbsession.query(TrackerField).get(int(id))
    if trackerfield:
        title = 'Edit Tracker Field'
    if request.method=='POST':
        if form.validate():
            if not trackerfield:
                trackerfield=TrackerField()
            form.populate_obj(trackerfield)
            trackerfield.tracker = tracker
            dbsession.add(trackerfield)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            return redirect('/trackers/view/' + str(trackerfield.tracker.id) + '#fields')
    else:
        if id:
            trackerfield = dbsession.query(TrackerField).get(int(id))
        if trackerfield:
            form = TrackerFieldForm(obj=trackerfield)
            title = 'Edit Tracker Field'
    return html(render(request,'generic/form.html',title=title,form=form,enctype='multipart/form-data'))

@bp.route('/trackers/deletefield/<field_id>',methods=['POST'])
@authorized(require_admin=True)
async def deletefield(request,field_id=None):
    if field_id:
        trackerfield = dbsession.query(TrackerField).get(int(field_id))
        trackerid = trackerfield.tracker.id
        if(trackerfield):
            dbsession.delete(trackerfield)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
        return redirect('/trackers/view/' + str(trackerid) + '#fields')
    else:
        request['session']['flashmessage']='You need to specify the field to delete'
        return redirect(request.app.url_for('trackers.index'))

@bp.route('/trackers/<module>/<slug>/field/<field_id>/json')
async def fieldjson(request,module,slug=None,field_id=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    if field_id:
        trackerfield = dbsession.query(TrackerField).get(int(field_id))
        if(trackerfield):
            if trackerfield.field_type=='user':
                sqlq = "select id,name from users where name ilike '%" + request.args['q'][0] + "%' "
            elif trackerfield.field_type=='object':
                sqlq = "select id," + trackerfield.main_obj_field() + " from " + trackerfield.obj_table + " where " + " or ".join([field + " ilike '%" + request.args['q'][0] + "%' " for field in trackerfield.obj_fields() ])
            results = dbsession.execute(sqlq)
            return jsonresponse([ {'id':result.id,'name':result.name} for result in results ])


@bp.route('/trackers/view/')
@bp.route('/trackers/view/<id>')
@authorized(require_admin=True)
async def view(request,id=None):
    if id:
        tracker = dbsession.query(Tracker).get(int(id))
        dataupdates = dbsession.query(TrackerDataUpdate).filter(TrackerDataUpdate.tracker==tracker)
        updatespaginator = Paginator(dataupdates, 10)
        return html(render(request,'trackers/view.html',tracker=tracker,updatespaginator=updatespaginator,curupdatepage=updatespaginator.page(int(request.args['updatepage'][0]) if 'updatepage' in request.args else 1)))
    else:
        return redirect('/')

@bp.route('/trackers/updatedb/<id>')
@authorized(require_admin=True)
async def updatedb(request,id=None):
    tracker = dbsession.query(Tracker).get(int(id))
    tracker.updatedb()
    return redirect('/trackers/view/' + str(id))

@bp.route('/trackers/create',methods=['POST','GET'],name='create')
@bp.route('/trackers/edit/',methods=['POST','GET'],name='edit')
@bp.route('/trackers/edit/<id>',methods=['POST','GET'],name='edit')
@authorized(require_admin=True)
async def form(request,id=None):
    title = 'Create Tracker'
    form = TrackerForm(request.form)
    tracker = None
    tokeninput = None
    newtracker = False
    if request.method=='POST':
        if id:
            tracker = dbsession.query(Tracker).get(int(id))
        if tracker:
            title = 'Edit Tracker'
        if form.validate():
            if not tracker:
                tracker=Tracker()
                newtracker = True
            form.populate_obj(tracker)
            if form['list_fields'].data:
                tracker.list_fields = ','.join([ ddat.name for ddat in dbsession.execute("select name from tracker_fields where id in (" + form['list_fields'].data + ")") ])
            if form['search_fields'].data:
                tracker.search_fields = ','.join([ ddat.name for ddat in dbsession.execute("select name from tracker_fields where id in (" + form['search_fields'].data + ")") ])
            if form['filter_fields'].data:
                tracker.filter_fields = ','.join([ ddat.name for ddat in dbsession.execute("select name from tracker_fields where id in (" + form['filter_fields'].data + ")") ])
            if form['excel_fields'].data:
                tracker.excel_fields = ','.join([ ddat.name for ddat in dbsession.execute("select name from tracker_fields where id in (" + form['excel_fields'].data + ")") ])
            dbsession.add(tracker)
            if newtracker:
                newstatus = TrackerStatus(name='New',tracker=tracker)
                dbsession.add(newstatus)
                adminrole = TrackerRole(name='Admin',role_type='module',tracker=tracker)
                dbsession.add(adminrole)
                statusfield = TrackerField(name='record_status',label='Status',tracker=tracker,field_type='string')
                dbsession.add(statusfield)
                idfield = TrackerField(name='id',label='ID',tracker=tracker,field_type='integer')
                dbsession.add(idfield)
                newtransition = TrackerTransition(name='new',tracker=tracker,roles=[adminrole],next_status=newstatus)
                dbsession.add(newtransition)
            success = False
            try:
                dbsession.commit()
                success = True
            except IntegrityError as inst:
                form.slug.errors.append('Tracker with slug ' + form.slug.data + ' already exist in module ' + form.module.data + '. It needs to be unique')
                dbsession.rollback()
            except Exception as inst:
                dbsession.rollback()
            if success:
                return redirect('/trackers/view/' + str(tracker.id))
    else:
        if id:
            tracker = dbsession.query(Tracker).get(int(id))
            if tracker:
                form = TrackerForm(obj=tracker)
                title = 'Edit Tracker'
                tokeninput = {
                        'list_fields': {
                            'url': request.app.url_for('trackers.trackerfieldsjson',module=tracker.slug,slug=tracker.slug),
                            'prePopulate':[ {'id':field.id,'name':field.name} for field in tracker.list_fields_list() ]
                            },
                        'search_fields': {
                            'url': request.app.url_for('trackers.trackerfieldsjson',module=tracker.module,slug=tracker.slug),
                            'prePopulate':[ {'id':field.id,'name':field.name} for field in tracker.fields_from_list(tracker.search_fields) ]
                            },
                        'filter_fields': {
                            'url': request.app.url_for('trackers.trackerfieldsjson',module=tracker.module,slug=tracker.slug),
                            'prePopulate':[ {'id':field.id,'name':field.name} for field in tracker.fields_from_list(tracker.filter_fields) ]
                            },
                        'excel_fields': {
                            'url': request.app.url_for('trackers.trackerfieldsjson',module=tracker.module,slug=tracker.slug),
                            'prePopulate':[ {'id':field.id,'name':field.name} for field in tracker.fields_from_list(tracker.excel_fields) ]
                            },
                        }

    return html(render(request,'generic/form.html',title=title,form=form,enctype='multipart/form-data',tokeninput=tokeninput))

@bp.route('/trackers')
@authorized(require_admin=True)
async def index(request):
    trackers = dbsession.query(Tracker)
    paginator = Paginator(trackers, 5)
    return html(render(request,
    'generic/list.html',title='Trackers',editlink='trackers.view',addlink='trackers.create',maxlength=100,fields=[{'label':'Module','name':'module'},{'label':'Slug','name':'slug'},{'label':'Title','name':'title'},{'label':'List Fields','name':'list_fields'},{'label':'Require Login','name':'require_login'},{'label':'Allowed Roles','name':'allowed_roles'}],paginator=paginator,curpage=paginator.page(int(request.args['page'][0]) if 'page' in request.args else 1)))

@bp.route('/trackers/delete/<module>/<slug>',methods=['POST'])
@authorized(require_admin=True)
async def delete(request,module,slug=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    for update in tracker.dataupdates:
        if update.filename and os.path.exists(update.filename):
            os.remove(update.filename)
        if os.path.exists(os.path.join(uploadfolder,tracker.slug,'dataupdate',str(update.id))):
            os.rmdir(os.path.join(uploadfolder,tracker.slug,'dataupdate',str(update.id)))
        dbsession.delete(update)
        try:
            dbsession.commit()
        except Exception as inst:
            dbsession.rollback()
    for trackerstatus in tracker.statuses:
        print('deleting status' + str(trackerstatus))
        dbsession.delete(trackerstatus)
    for trackerrole in tracker.roles:
        print('deleting role' + str(trackerrole))
        dbsession.delete(trackerrole)
    for trackerfield in tracker.fields:
        print('deleting field' + str(trackerfield))
        dbsession.delete(trackerfield)
    for trackertransition in tracker.transitions:
        print('deleting transition' + str(trackertransition))
        dbsession.delete(trackertransition)
    dbsession.execute("drop table if exists " + tracker.data_table() + ";")
    dbsession.execute("drop sequence if exists public." + tracker.data_table() + "_id_seq;")
    dbsession.execute("drop table if exists " + tracker.update_table() + ";")
    dbsession.execute("drop sequence if exists public." + tracker.update_table() + "_id_seq;")
    dbsession.delete(tracker)
    try:
        dbsession.commit()
    except Exception as inst:
        dbsession.rollback()
    return redirect('/trackers/')

@bp.route('/system/<module>/<slug>/')
@authorized(object_type='tracker')
async def viewlist(request,module,slug=None):
    if slug==None:
        slug=module
        module='portal'
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    page = dbsession.query(Page).filter_by(slug=tracker.slug + '_list').first()
    if page:
        return html(page.render(request,tracker=tracker))
    else:
        return html(render(request,'trackers/viewlist.html',tracker=tracker))

@bp.route('/system/<module>/<slug>/excel')
@authorized(object_type='tracker')
async def listexcel(request,module,slug=None):
    if slug==None:
        slug=module
        module='portal'
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    curuser = None
    if 'user_id' in request['session']:
        curuser = dbsession.query(User).filter(User.id==request['session']['user_id']).first()
    print("rcd:" + str(request))
    records = tracker.records(curuser=curuser,request=request)
    wb = Workbook()
    ws = wb.active
    ws.title = tracker.title
    if records:
        xfields = []
        if tracker.excel_fields:
            xfields = tracker.fields_from_list(tracker.excel_fields)
        else:
            xfields = tracker.fields_from_list(tracker.list_fields)
        for i,f in enumerate(xfields):
            ws.cell(row=1,column=i+1,value=f.label)
        row=2
        for rec in records:
            for i,f in enumerate(xfields):
                ws.cell(row=row,column=i+1,value=rec[f.name])
            row+=1

    virtual_wb = save_virtual_workbook(wb)
    return raw(virtual_wb, content_type='application/vnd.ms-excel')

@bp.route('/system/<module>/<slug>/addrecord',methods=['POST','GET'])
@authorized(object_type='tracker')
async def addrecord(request,module,slug=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    newtransition = dbsession.query(TrackerTransition).filter_by(tracker=tracker,name='new').first()
    if request.method=='POST':
        data = tracker.addrecord(request.form,request)
        if newtransition.postpage:
            output=None
            ldict = locals()
            exec(newtransition.postpage,globals(),ldict)
            if 'output' in ldict and str(ldict['output']):
                return redirect(str(ldict['output']))
            else:
                return redirect(request.app.url_for('trackers.viewrecord',slug=tracker.slug,id=data['id']))
        else:
            return redirect(request.app.url_for('trackers.viewrecord',slug=tracker.slug,id=data['id']))
    page = dbsession.query(Page).filter_by(slug=tracker.slug + '_addrecord').first()
    if page:
        return html(page.render(request,tracker=tracker,transition=newtransition))
    else:
        return html(render(request,'trackers/formrecord.html',tracker=tracker,transition=newtransition))

@bp.route('/system/<module>/<slug>/edit/<transition_id>/<record_id>',methods=['POST','GET'])
@authorized(object_type='tracker')
async def editrecord(request,module,slug=None,transition_id=None,record_id=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    transition = None
    record = None
    curuser = None
    if 'user_id' in request['session']:
        curuser = dbsession.query(User).filter(User.id==request['session']['user_id']).first()
    if record_id:
        record = tracker.records(record_id,curuser=curuser,request=request)
    if transition_id:
        transition = dbsession.query(TrackerTransition).get(transition_id)
    if request.method=='POST':
        data = tracker.editrecord(request.form,request,id=record_id)
        if transition.postpage:
            output=None
            ldict = locals()
            exec(transition.postpage,globals(),ldict)
            if 'output' in ldict and str(ldict['output']):
                return redirect(str(ldict['output']))
            else:
                return redirect(request.app.url_for('trackers.viewrecord',slug=tracker.slug,id=data['id']))
        else:
            return redirect(request.app.url_for('trackers.viewrecord',slug=tracker.slug,id=record_id))
    page = dbsession.query(Page).filter_by(slug=tracker.slug + '_editrecord').first()
    if page:
        return html(page.render(request,tracker=tracker,transition=transition,record=record))
    else:
        return html(render(request,'trackers/formrecord.html',tracker=tracker,transition=transition,record=record))

@bp.route('/system/<module>/<slug>/<id>',methods=['POST','GET'])
@authorized(object_type='tracker')
async def viewrecord(request,module,slug=None,id=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    curuser = None
    if 'user_id' in request['session']:
        curuser = dbsession.query(User).filter(User.id==request['session']['user_id']).first()
    record = None
    status = None
    page = None
    if request.method=='POST':
        return redirect('/system/' + slug + '/viewrecord/' + id)
    if(id):
        record = tracker.records(id,curuser=curuser,request=request)
        if record:
            status = tracker.status(record)
    if status:
        page = dbsession.query(Page).filter_by(slug=tracker.slug + '_view_' + status.name.lower().replace(' ','_')).first()
    if not page:
        page = dbsession.query(Page).filter_by(slug=tracker.slug + '_view_default').first()
    if page:
        return html(page.render(request,tracker=tracker,record=record))
    else:
        return html(render(request,'trackers/viewrecord.html',tracker=tracker,record=record))
