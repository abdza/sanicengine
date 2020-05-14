# -*- coding: utf-8 -*-
from sanic import Blueprint
from sanic.response import html, redirect, json as jsonresponse, file_stream, stream, raw
from .models import Tracker, TrackerField, TrackerRole, TrackerStatus, TrackerTransition, TrackerDataUpdate
from sanicengine.trees.models import Tree
from sanicengine.users.models import User
from sanicengine.pages.models import Page
from sanicengine.emailtemplates.models import EmailTemplate
from .forms import TrackerForm, TrackerFieldForm, TrackerRoleForm, TrackerStatusForm, TrackerTransitionForm
from sanicengine.database import dbsession, executedb
from sanicengine.template import render, render_string
from sanicengine.decorators import authorized
from sqlalchemy_paginator import Paginator
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError

import os
import datetime
import json
import re
import asyncio
import io
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook


bp = Blueprint('trackers')

uploadfolder = 'upload'

def slugify(slug):
    if(slug):
        slug = re.sub('[^0-9a-zA-Z]+', '_', slug.lower())
    return slug

@bp.route('/trackers/runupdate/<module>/<slug>',methods=['GET'])
@bp.route('/trackers/runupdate/<module>',methods=['GET'])
async def runupdate(request,module,slug=None):
    if slug==None:
        slug=module
        module='portal'
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    updates = dbsession.query(TrackerDataUpdate).filter_by(status='new').all()
    for update in updates:
        update.status = 'in queue'
        dbsession.add(update)
    try:
        dbsession.commit()
    except Exception as inst:
        dbsession.rollback()

    try:
        await asyncio.wait([ update.run() for update in updates])
    except:
        print("Nothing to wait for")
    return redirect(request.app.url_for('trackers.view',id=tracker.id) + '#dataupdates')

@bp.route('/trackers/fixstatus/<module>/<slug>',methods=['POST'])
@authorized(object_type='dataupdate')
async def fixstatus(request,module,slug):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    if tracker:

        newtransition = dbsession.query(TrackerTransition).filter(TrackerTransition.name.ilike("%new%"),TrackerTransition.tracker==tracker).first()

        if not newtransition:
            newtransition = TrackerTransition(tracker=tracker,name='New',label='New',edit_fields=tracker.list_fields,roles=[adminrole,],next_status=newstatus)
            dbsession.add(newtransition)
            tracker.default_new_transition = 'New'
            dbsession.add(tracker)

        statuses = []
        for status in tracker.statuses:
            statuses.append(status.name)
        
        fixquery = "select * from " + tracker.data_table + " where record_status not in ('" + "','".join(statuses) + "') or record_status is null"
        brokendatas = executedb(fixquery)
        for broken in brokendatas:
            tracker.saverecord({'id':broken.id,'record_status':newtransition.next_status})
        
    request['session']['flashmessage'] = 'Fixed record status for ' + tracker.title
    return redirect(request.app.url_for('trackers.view',id=tracker.id))

@bp.route('/trackers/cleardata/<module>/<slug>',methods=['POST'])
@authorized(object_type='dataupdate')
async def cleardata(request,module,slug):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    if tracker:
        for update in tracker.dataupdates:
            if update.filename and os.path.exists(update.filename):
                os.remove(update.filename)
            if os.path.exists(os.path.join(uploadfolder,tracker.slug,'dataupdate',str(update.id))):
                os.rmdir(os.path.join(uploadfolder,tracker.slug,'dataupdate',str(update.id)))
            dbsession.delete(update)
        try:
            dbsession.commit()
        except Exception as inst:
            dbsession.rollback()
        try:
            dbsession.execute("delete from " + tracker.update_table)
            dbsession.execute("delete from " + tracker.data_table)
        except:
            dbsession.rollback()
        request['session']['flashmessage'] = 'Cleared data for ' + tracker.title
    return redirect(request.app.url_for('trackers.view',id=tracker.id))


@bp.route('/trackers/deleteupdate/<update_id>',methods=['POST'])
@authorized(object_type='dataupdate')
async def deleteupdate(request,update_id=None):
    if update_id:
        update = dbsession.query(TrackerDataUpdate).get(int(update_id))
        trackerid = update.tracker.id
        if(update):
            if update.filename and os.path.exists(update.filename):
                os.remove(update.filename)
            if os.path.exists(os.path.join(uploadfolder,update.tracker.slug,'dataupdate',str(update.id))):
                os.rmdir(os.path.join(uploadfolder,update.tracker.slug,'dataupdate',str(update.id)))
            try:
                dbsession.execute("delete from " + update.tracker.update_table + " where record_id in (select id from " + update.tracker.data_table + " where batch_no=" + str(update.id) + ")")
                dbsession.execute("delete from " + update.tracker.data_table + " where batch_no=" + str(update.id))
            except:
                dbsession.rollback()

            dbsession.delete(update)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
        return redirect(request.app.url_for('trackers.view',id=trackerid) + '#dataupdates')
    else:
        request['session']['flashmessage']='Need to specify id of update to delete'
        return redirect(request.app.url_for('trackers.index'))

@bp.route('/trackers/dataupdate/<module>/<slug>',methods=['GET','POST'])
@bp.route('/trackers/dataupdate/<module>',methods=['GET','POST'])
@authorized(object_type='dataupdate')
async def data_update(request,module,slug=None):
    if slug==None:
        slug=module
        module='portal'
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    columns = []
    dataupdate = None
    if request.method=='POST':
        if request.form.get('dataupdate_id'):
            dataupdate = dbsession.query(TrackerDataUpdate).get(str(request.form.get('dataupdate_id')))
            dataupdate.data_params=json.dumps(request.form)
            dbsession.add(dataupdate)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            return redirect(request.app.url_for('trackers.view',id=tracker.id) + '#dataupdates')
        if request.files.get('excelfile'):
            dataupdate = TrackerDataUpdate(status='new',tracker=tracker,created_date=datetime.datetime.now())
            dbsession.add(dataupdate)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            dfile = request.files.get('excelfile')
            ext = dfile.type.split('/')[1]
            if not os.path.exists(os.path.join(uploadfolder,tracker.slug,'dataupdate',str(dataupdate.id))):
                os.makedirs(os.path.join(uploadfolder,tracker.slug,'dataupdate',str(dataupdate.id)))
            dst = os.path.join(uploadfolder,tracker.slug,'dataupdate',str(dataupdate.id),dfile.name)
            try:
                # extract starting byte from Content-Range header string
                range_str = request.headers['Content-Range']
                start_bytes = int(range_str.split(' ')[1].split('-')[0])
                with open(dst, 'ab') as f:
                    f.seek(start_bytes)
                    f.write(dfile.body)
            except KeyError:
                with open(dst, 'wb') as f:
                    f.write(dfile.body)
            dataupdate.filename = dst
            dbsession.add(dataupdate)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            wb = load_workbook(filename = dst)
            ws = wb.active
            columns.append({'field_val':'ignore','field_name':'Ignore'})
            columns.append({'field_val':'custom','field_name':'Custom'})
            for row in ws.iter_rows(max_row=1):
                for cell in row:
                    if cell.value and cell.value.lower() != 'id':
                        columns.append({'field_val':cell.column,'field_name':cell.value})
    return html(render(request,'/trackers/data_update.html',tracker=tracker,columns=columns,dataupdate=dataupdate),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})

@bp.route('/trackers/create_from_excel/<module>/<slug>',methods=['POST','GET'])
@bp.route('/trackers/create_from_excel/<module>',methods=['POST','GET'])
@authorized(require_admin=True)
async def create_from_excel(request,module,slug=None):
    if slug==None:
        slug=module
        module='portal'
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    fields = []
    field_types = []
    if tracker and request.method=='POST':
        if request.form.get('field_name'):
            field_names = request.form['field_name']
            field_label = request.form['field_label']
            field_type = request.form['field_type']
            list_fields = []
            for i,name in enumerate(field_names):
                if not field_type[i]=='ignore':
                    tfield = TrackerField(tracker=tracker, name=name, label=field_label[i], field_type=field_type[i])
                    dbsession.add(tfield)
                    list_fields.append(name)
            tracker.list_fields = ','.join(list_fields)
            tracker.detail_fields = ','.join(list_fields)
            dbsession.add(tracker)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            return redirect(request.app.url_for('trackers.view',id=tracker.id) + '#fields')
        if request.files.get('excelfile'):
            dfile = request.files.get('excelfile')
            ext = dfile.type.split('/')[1]
            dst = os.path.join(uploadfolder,dfile.name)
            if not os.path.exists(uploadfolder):
                os.makedirs(uploadfolder)
            try:
                # extract starting byte from Content-Range header string
                range_str = request.headers['Content-Range']
                start_bytes = int(range_str.split(' ')[1].split('-')[0])
                with open(dst, 'ab') as f:
                    f.seek(start_bytes)
                    f.write(dfile.body)
            except KeyError:
                with open(dst, 'wb') as f:
                    f.write(dfile.body)
            wb = load_workbook(filename = dst)
            ws = wb.active
            fieldtitles = []
            fieldtypes = []
            for row in ws.iter_rows(max_row=1):
                for cell in row:
                    if cell.value!=None:
                        fieldtitles.append(cell.value)
            for row in ws.iter_rows(max_row=2,min_row=2):
                currow = 0
                for cell in row:
                    if cell.value!=None and currow<len(fieldtitles):
                        fieldtypes.append(cell.data_type)
                    else:
                        fieldtypes.append('string')
                    currow += 1
            for i,title in enumerate(fieldtitles):
                if slugify(title)!='id':
                    fields.append({'field_name':slugify(title),'field_type':fieldtypes[i]})
            os.remove(dst)
        field_types = [('ignore','Ignore'),('string','String'),('text','Text'),('integer','Integer'),('number','Number'),('date','Date'),('datetime','Date Time'),('boolean','Boolean'),('object','Object'),('treenode','TreeNode')]
    return html(render(request,'/trackers/create_from_excel.html',tracker=tracker,fields=fields,field_types=field_types),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})

@bp.route('/trackers/editstatus/<module>/<slug>/<id>',methods=['POST','GET'])
@bp.route('/trackers/editstatus/<module>/<slug>/',methods=['POST','GET'],name='editstatus')
@bp.route('/trackers/addstatus/<module>/<slug>',methods=['POST','GET'],name='createstatus')
@authorized(require_admin=True)
async def statusform(request,module,slug,id=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    title = 'Create Tracker Status'
    form = TrackerStatusForm(request.form)
    trackerstatus = None
    if id:
        trackerstatus = dbsession.query(TrackerStatus).get(int(id))
    if trackerstatus:
        title = 'Edit Tracker Status'
    tokeninput = {
            'display_fields': {
                'url': request.app.url_for('trackers.trackerfieldsjson',module=module,slug=slug),
                },
            }
    if request.method=='POST':
        if form.validate():
            if not trackerstatus:
                trackerstatus=TrackerStatus()
            form.populate_obj(trackerstatus)
            trackerstatus.tracker = tracker
            if form['display_fields'].data:
                trackerstatus.display_fields = ','.join(TrackerField.listnames(form['display_fields'].data))

            dbsession.add(trackerstatus)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            return redirect(request.app.url_for('trackers.view',id=trackerstatus.tracker.id) + '#statuses')
    else:
        if id:
            trackerstatus = dbsession.query(TrackerStatus).get(int(id))
        if trackerstatus:
            form = TrackerStatusForm(obj=trackerstatus)
            title = 'Edit Tracker Status'
            tokeninput = {
                    'display_fields': {
                        'url': request.app.url_for('trackers.trackerfieldsjson',module=trackerstatus.tracker.module,slug=trackerstatus.tracker.slug),
                        'prePopulate':[ {'id':field.id,'name':field.name} for field in trackerstatus.tracker.fields_from_list(trackerstatus.display_fields) ]
                        },
                    }
    return html(render(request,'generic/form.html',title=title,form=form,enctype='multipart/form-data',tokeninput=tokeninput),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})

@bp.route('/trackers/deletestatus/<status_id>',methods=['POST'])
@authorized(require_admin=True)
async def deletestatus(request,status_id=None):
    if status_id:
        trackerstatus = dbsession.query(TrackerStatus).get(int(status_id))
        trackerid = trackerstatus.tracker.id
        if(trackerstatus):
            dbsession.delete(trackerstatus)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
        return redirect(request.app.url_for('trackers.view',id=trackerid) + '#statuses')
    else:
        request['session']['flashmessage']='Need to specify id of status to delete'
        redirect(request.app.url_for('trackers.index'))

@bp.route('/trackers/roles/<module>/<slug>/json')
async def rolesjson(request,module,slug=None):
    if slug==None:
        slug=module
        module='portal'
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    trackerroles = dbsession.query(TrackerRole).filter(TrackerRole.tracker==tracker,TrackerRole.name.ilike('%' + request.args['q'][0] + '%')).all() 
    return jsonresponse([ {'id':role.id,'name':role.name} for role in trackerroles ])

@bp.route('/trackers/fields/<module>/<slug>/json')
async def trackerfieldsjson(request,module,slug=None):
    if slug==None:
        slug=module
        module='portal'
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    trackerfields = dbsession.query(TrackerField).filter(TrackerField.tracker==tracker,TrackerField.name.ilike('%' + request.args['q'][0] + '%')).all() 
    return jsonresponse([ {'id':field.id,'name':field.name} for field in trackerfields ])


@bp.route('/trackers/transition/<module>/<slug>/edit/<id>',methods=['POST','GET'])
@bp.route('/trackers/transition/<module>/<slug>/edit/',methods=['POST','GET'],name='edittransition')
@bp.route('/trackers/transition/<module>/<slug>/add',methods=['POST','GET'],name='createtransition')
@authorized(require_admin=True)
async def transitionform(request,module,slug=None,id=None):

    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    title = tracker.title + '-Create Transition'
    form = TrackerTransitionForm(request.form)
    dstatuses = [('','None'),] + [(str(g.id),g.name) for g in dbsession.query(TrackerStatus).filter(TrackerStatus.tracker==tracker).all()]
    form.prev_status_id.choices = dstatuses
    form.next_status_id.choices = dstatuses
    trackertransition = None
    tokeninput = {
            'emails': {
                'url': request.app.url_for('emailtemplates.jsonlist',module=module),
                },
            'display_fields': {
                'url': request.app.url_for('trackers.trackerfieldsjson',module=module,slug=slug),
                },
            'edit_fields': {
                'url': request.app.url_for('trackers.trackerfieldsjson',module=module,slug=slug),
                },
            'roles': {
                'url': request.app.url_for('trackers.rolesjson',module=module,slug=slug),
                },
            }
    if id:
        trackertransition = dbsession.query(TrackerTransition).get(int(id))
    if trackertransition:
        title = tracker.title + '-Edit Transition'
    if request.method=='POST':
        if form.validate():
            if not trackertransition:
                trackertransition=TrackerTransition()
            if form.roles.data:
                role_ids = form.roles.data.split(',')
                roles = [ dbsession.query(TrackerRole).get(role_id) for role_id in role_ids ]
                trackertransition.roles = roles
            else:
                trackertransition.roles = []
            del(form['roles'])
            if form['emails'].data:
                email_ids = form['emails'].data.split(',')
                emails = [ dbsession.query(EmailTemplate).get(email_id) for email_id in email_ids ]
                trackertransition.emails = emails
            else:
                trackertransition.emails = []
            del(form['emails'])
            display_fields = []
            edit_fields = []
            detail_fields = []
            if form['display_fields'].data:
                display_fields = ','.join(TrackerField.listnames(form['display_fields'].data))
                del(form['display_fields'])
            if form['edit_fields'].data:
                edit_fields = ','.join(TrackerField.listnames(form['edit_fields'].data))
                del(form['edit_fields'])
            form.populate_obj(trackertransition)
            trackertransition.display_fields = display_fields
            trackertransition.edit_fields = edit_fields
            trackertransition.detail_fields = detail_fields
            if form.prev_status_id.data=='':
                trackertransition.prev_status = None
                trackertransition.prev_status_id = None
                del(form['prev_status_id'])
            if form.next_status_id.data=='':
                trackertransition.next_status = None
                trackertransition.next_status_id = None
                del(form['next_status_id'])
            form.populate_obj(trackertransition)
            trackertransition.tracker = tracker
            dbsession.add(trackertransition)
            try:
                dbsession.commit()
                for eid in emails:
                    curemail = dbsession.query(EmailTemplate).get(int(eid.strip()))
                    if curemail:
                        curemail.transition_id = trackertransition.id
                        dbsession.add(curemail)
                        dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            return redirect(request.app.url_for('trackers.view',id=trackertransition.tracker.id) + '#transitions')
    else:
        if id:
            trackertransition = dbsession.query(TrackerTransition).get(int(id))
        if trackertransition:
            form = TrackerTransitionForm(obj=trackertransition)
            form.prev_status_id.choices = dstatuses
            form.next_status_id.choices = dstatuses
            title = tracker.title + '-Edit Transition'
            tokeninput['display_fields']['prePopulate'] = [ {'id':field.id,'name':field.name} for field in trackertransition.tracker.fields_from_list(trackertransition.display_fields) ]
            tokeninput['emails']['prePopulate'] = [ {'id':field.id,'name':field.title} for field in trackertransition.emails ]
            tokeninput['edit_fields']['prePopulate'] = [ {'id':field.id,'name':field.name} for field in trackertransition.tracker.fields_from_list(trackertransition.edit_fields) ]
            tokeninput['roles']['prePopulate'] = [ {'id':field.id,'name':field.name} for field in trackertransition.roles ]
    return html(render(request,'generic/form.html',title=title,form=form,tracker=tracker,enctype='multipart/form-data',tokeninput=tokeninput,acefield=['postpage'],acetype={'postpage':'python'}),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})

@bp.route('/trackers/deletetransition/<transition_id>',methods=['POST'])
@authorized(require_admin=True)
async def deletetransition(request,transition_id=None):
    if transition_id:
        trackertransition = dbsession.query(TrackerTransition).get(int(transition_id))
        trackerid=trackertransition.tracker.id
        if(trackertransition):
            dbsession.delete(trackertransition)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
        return redirect(request.app.url_for('trackers.view', id=trackerid) + '#transitions')
    else:
        request['session']['flashmessage']='You need to specify id of transition to delete'
        return redirect(request.app.url_for('trackers.index'))

@bp.route('/trackers/role/<module>/<slug>/edit/<id>',methods=['POST','GET'])
@bp.route('/trackers/role/<module>/<slug>/edit/',methods=['POST','GET'],name='editrole')
@bp.route('/trackers/role/<module>/<slug>/add',methods=['POST','GET'],name='createrole')
@authorized(require_admin=True)
async def roleform(request,module,slug=None,id=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    title = tracker.title + '-Create Role'
    form = TrackerRoleForm(request.form)
    trackerrole = None
    if id:
        trackerrole = dbsession.query(TrackerRole).get(int(id))
    if trackerrole:
        title = tracker.title + '-Edit Role'
    if request.method=='POST':
        if form.validate():
            if not trackerrole:
                trackerrole=TrackerRole()
            form.populate_obj(trackerrole)
            trackerrole.tracker = tracker
            dbsession.add(trackerrole)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            return redirect(request.app.url_for('trackers.view',id=trackerrole.tracker.id) + '#roles')
    else:
        if id:
            trackerrole = dbsession.query(TrackerRole).get(int(id))
        if trackerrole:
            form = TrackerRoleForm(obj=trackerrole)
            title = tracker.title + '-Edit Role'
    return html(render(request,'generic/form.html',title=title,form=form,enctype='multipart/form-data',acefield=['compare'],acetype={'compare':'python'}),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})

@bp.route('/trackers/deleterole/<role_id>',methods=['POST'])
@authorized(require_admin=True)
async def deleterole(request,role_id=None):
    if role_id:
        trackerrole = dbsession.query(TrackerRole).get(int(role_id))
        trackerid = trackerrole.tracker.id
        if(trackerrole):
            dbsession.delete(trackerrole)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
        return redirect(request.app.url_for('trackers.view',id=trackerid) + '#roles')
    else:
        request['session']['flashmessage']='You need to specify id of role to delete'
        return redirect(request.app.url_for('trackers.index'))

@bp.route('/trackers/field/<module>/<slug>/edit/<id>',methods=['POST','GET'])
@bp.route('/trackers/field/<module>/<slug>/edit/',methods=['POST','GET'],name='editfield')
@bp.route('/trackers/field/<module>/<slug>/add',methods=['POST','GET'],name='createfield')
@authorized(require_admin=True)
async def fieldform(request,module,slug=None,id=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    title = tracker.title + '-Create Field'
    form = TrackerFieldForm(request.form)
    trackerfield = None
    if id:
        trackerfield = dbsession.query(TrackerField).get(int(id))
    if trackerfield:
        title = tracker.title + '-Edit Field'
    if request.method=='POST':
        if form.validate():
            if not trackerfield:
                trackerfield=TrackerField()
            form.populate_obj(trackerfield)
            trackerfield.tracker = tracker
            dbsession.add(trackerfield)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
            return redirect(request.app.url_for('trackers.view',id=trackerfield.tracker.id) + '#fields')
    else:
        if id:
            trackerfield = dbsession.query(TrackerField).get(int(id))
        if trackerfield:
            form = TrackerFieldForm(obj=trackerfield)
            title = tracker.title + '-Edit Field'
    return html(render(request,'generic/form.html',title=title,form=form,enctype='multipart/form-data',acefield=['default','options'],acetype={'default':'python','options':'python'},info="<p>Default will take final value of the output variable. curuser is available if the user is logged in. The form variable is also available to check on the information submitted in other fields. If type is boolean, then options can be value when false/true in the form of '&lt;false value&gt;,&lt;true value&gt;' </p>"),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})

@bp.route('/trackers/deletefield/<field_id>',methods=['POST'])
@authorized(require_admin=True)
async def deletefield(request,field_id=None):
    if field_id:
        trackerfield = dbsession.query(TrackerField).get(int(field_id))
        trackerid = trackerfield.tracker.id
        if(trackerfield):
            dbsession.delete(trackerfield)
            try:
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
        return redirect(request.app.url_for('trackers.view',id=trackerid) + '#fields')
    else:
        request['session']['flashmessage']='You need to specify the field to delete'
        return redirect(request.app.url_for('trackers.index'))

@bp.route('/trackers/field/<module>/<slug>/json/<field_id>')
async def fieldjson(request,module,slug=None,field_id=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    if field_id:
        trackerfield = dbsession.query(TrackerField).get(int(field_id))
        if(trackerfield):
            namefield = "name"
            if trackerfield.field_type=='user':
                namefield = "name"
                sqlq = "select id,name from users where name ilike '%" + request.args['q'][0] + "%' "
            elif trackerfield.field_type=='object':
                namefield = trackerfield.main_obj_field()
                qfield = None
                if trackerfield.obj_filter:
                    try:
                        qfield = render_string(request,trackerfield.obj_filter)
                    except:
                        print("Error rendering query filter:" + str(trackerfield.obj_filter))
                sqlq = "select id," + trackerfield.main_obj_field() + " from " + trackerfield.obj_table + " where (" + " or ".join([field + " ilike '%" + request.args['q'][0] + "%' " for field in trackerfield.obj_fields() ]) + ") " + (" and " + qfield if qfield else "")
            elif trackerfield.field_type=='treenode':
                namefield = "titleagg"
                tmodule = trackerfield.tracker.module
                tslug = trackerfield.tracker.slug
                tparts = trackerfield.obj_table.split('.')
                if len(tparts)==2:
                    tmodule = tparts[0]
                    tslug = tparts[1]
                tree = dbsession.query(Tree).filter(Tree.module==tmodule,Tree.slug==tslug).first()
                if tree:
                    sqlq = "select id,titleagg from (select nleaf.id,nleaf.lft,nleaf.rgt,string_agg(cnode.title,'->' order by cnode.lft) titleagg from tree_nodes cnode,(select id,lft,rgt,tree_id from tree_nodes where rgt-lft=1 and tree_id=" + str(tree.id) + ") nleaf where cnode.lft<=nleaf.lft and cnode.rgt>=nleaf.rgt and cnode.tree_id=nleaf.tree_id group by nleaf.lft, nleaf.rgt, nleaf.id) combtitle where titleagg ilike '%" + "%".join(request.args['q'][0].replace(" ","")) + "%' limit 20"
                else:
                    sqlq = "select id,titleagg from (select nleaf.id,nleaf.lft,nleaf.rgt,string_agg(cnode.title,'->' order by cnode.lft) titleagg from tree_nodes cnode,(select id,lft,rgt,tree_id from tree_nodes where rgt-lft=1) nleaf where cnode.lft<=nleaf.lft and cnode.rgt>=nleaf.rgt and cnode.tree_id=nleaf.tree_id group by nleaf.lft, nleaf.rgt, nleaf.id) combtitle where titleagg ilike '%" + "%".join(request.args['q'][0].replace(" ","")) + "%' limit 20"
            results = dbsession.execute(sqlq)
            return jsonresponse([ {'id':result.id,'name':result[namefield]} for result in results ])


@bp.route('/trackers/view/<id>')
@bp.route('/trackers/view/',name='viewtracker')
@authorized(require_admin=True)
async def view(request,id=None):
    if id:
        tracker = dbsession.query(Tracker).get(int(id))
        dataupdates = dbsession.query(TrackerDataUpdate).filter(TrackerDataUpdate.tracker==tracker)
        updatespaginator = Paginator(dataupdates, 10)
        return html(render(request,'trackers/view.html',title=tracker.title,tracker=tracker,updatespaginator=updatespaginator,curupdatepage=updatespaginator.page(int(request.args['updatepage'][0]) if 'updatepage' in request.args else 1)),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})
    else:
        return redirect('/')

@bp.route('/trackers/createpages/<id>')
@authorized(require_admin=True)
async def createpages(request,id=None):
    tracker = dbsession.query(Tracker).get(int(id))
    page = dbsession.query(Page).filter_by(module=tracker.module,slug=tracker.slug + '_list').first()
    if not page:
        page = Page(title=tracker.title + ' List',module=tracker.module,slug=tracker.slug + '_list',content=open('sanicengine/templates/trackers/viewlist.html').read())
        dbsession.add(page)
    page = dbsession.query(Page).filter_by(module=tracker.module,slug=tracker.slug + '_addrecord').first()
    if not page:
        page = Page(title=tracker.title + ' Add Record',module=tracker.module,slug=tracker.slug + '_addrecord',content=open('sanicengine/templates/trackers/formrecord.html').read())
        dbsession.add(page)
    page = dbsession.query(Page).filter_by(module=tracker.module,slug=tracker.slug + '_editrecord').first()
    if not page:
        page = Page(title=tracker.title + ' Edit Record',module=tracker.module,slug=tracker.slug + '_editrecord',content=open('sanicengine/templates/trackers/formrecord.html').read())
        dbsession.add(page)
    page = dbsession.query(Page).filter_by(module=tracker.module,slug=tracker.slug + '_view_default').first()
    if not page:
        page = Page(title=tracker.title + ' View',module=tracker.module,slug=tracker.slug + '_view_default',content=open('sanicengine/templates/trackers/viewrecord.html').read())
        dbsession.add(page)

    dbsession.commit()

    return redirect(request.app.url_for('trackers.view',id=id))

@bp.route('/trackers/updatedb/<id>')
@authorized(require_admin=True)
async def updatedb(request,id=None):
    tracker = dbsession.query(Tracker).get(int(id))
    tracker.updatedb()
    return redirect(request.app.url_for('trackers.view',id=id))

@bp.route('/trackers/defaulttransitions/<id>')
@authorized(require_admin=True)
async def defaulttransitions(request,id=None):
    tracker = dbsession.query(Tracker).get(int(id))
    if tracker and tracker.list_fields:

        if len(tracker.detail_fields)==0:
            tracker.detail_fields = tracker.list_fields
            dbsession.add(tracker)

        newstatus = dbsession.query(TrackerStatus).filter(TrackerStatus.name.ilike("%new%"),TrackerStatus.tracker==tracker).first()
        adminrole = dbsession.query(TrackerRole).filter(TrackerRole.name.ilike("%admin%"),TrackerRole.tracker==tracker).first()

        newtransition = dbsession.query(TrackerTransition).filter(TrackerTransition.name.ilike("%new%"),TrackerTransition.tracker==tracker).first()

        if not newtransition:
            newtransition = TrackerTransition(tracker=tracker,name='New',label='New',edit_fields=tracker.list_fields,roles=[adminrole,],next_status=newstatus)
            dbsession.add(newtransition)
            tracker.default_new_transition = 'New'
            dbsession.add(tracker)

        edittransition = dbsession.query(TrackerTransition).filter(TrackerTransition.name.ilike("%edit%"),TrackerTransition.tracker==tracker).first()
        if not edittransition:
            edittransition = TrackerTransition(tracker=tracker,name='Edit',label='Edit',edit_fields=tracker.list_fields,roles=[adminrole,],prev_status=newstatus,next_status=newstatus)
            dbsession.add(edittransition)

        deletestatus = dbsession.query(TrackerStatus).filter(TrackerStatus.name.ilike("%delete%"),TrackerStatus.tracker==tracker).first()
        if not deletestatus:
            deletestatus = TrackerStatus(name='Delete',tracker=tracker)
            dbsession.add(deletestatus)

        deletetransition = dbsession.query(TrackerTransition).filter(TrackerTransition.name.ilike("%delete%"),TrackerTransition.tracker==tracker).first()
        if not deletetransition:
            deletetransition = TrackerTransition(tracker=tracker,name='Delete',label='Delete',roles=[adminrole,],prev_status=newstatus,next_status=deletestatus)
            dbsession.add(deletetransition)

        try:
            dbsession.commit()
        except Exception as inst:
            dbsession.rollback()
    
    return redirect(request.app.url_for('trackers.view',id=id))

@bp.route('/trackers/edit/<id>',methods=['POST','GET'])
@bp.route('/trackers/edit',methods=['POST','GET'],name='edit')
@bp.route('/trackers/create',methods=['POST','GET'],name='create')
@authorized(require_admin=True)
async def form(request,id=None):
    title = 'Create Tracker'
    form = TrackerForm(request.form)
    tracker = None
    tokeninput = None
    newtracker = False
    if request.method=='POST':
        if id:
            tracker = dbsession.query(Tracker).get(int(id))
        if tracker:
            title = 'Edit Tracker'
        if form.validate():
            if not tracker:
                tracker=Tracker()
                newtracker = True
            form.populate_obj(tracker)
            if form['list_fields'].data:
                tracker.list_fields = ','.join(TrackerField.listnames(form['list_fields'].data))
            if form['search_fields'].data:
                tracker.search_fields = ','.join(TrackerField.listnames(form['search_fields'].data))
            if form['filter_fields'].data:
                tracker.filter_fields = ','.join(TrackerField.listnames(form['filter_fields'].data))
            if form['excel_fields'].data:
                tracker.excel_fields = ','.join(TrackerField.listnames(form['excel_fields'].data))
            if form['detail_fields'].data:
                tracker.detail_fields = ','.join(TrackerField.listnames(form['detail_fields'].data))
            dbsession.add(tracker)
            if newtracker:
                newstatus = TrackerStatus(name='New',tracker=tracker)
                dbsession.add(newstatus)
                adminrole = TrackerRole(name='Admin',role_type='module',tracker=tracker)
                dbsession.add(adminrole)
                statusfield = TrackerField(name='record_status',label='Status',tracker=tracker,field_type='string')
                dbsession.add(statusfield)
                idfield = TrackerField(name='id',label='ID',tracker=tracker,field_type='integer')
                dbsession.add(idfield)
            success = False
            try:
                dbsession.commit()
                success = True
            except IntegrityError as inst:
                form.slug.errors.append('Tracker with slug ' + form.slug.data + ' already exist in module ' + form.module.data + '. It needs to be unique')
                dbsession.rollback()
            except Exception as inst:
                dbsession.rollback()
            if success:
                return redirect(request.app.url_for('trackers.view',id=tracker.id))
    else:
        if id:
            tracker = dbsession.query(Tracker).get(int(id))
            if tracker:
                form = TrackerForm(obj=tracker)
                title = 'Edit Tracker'
                tokeninput = {
                        'list_fields': {
                            'url': request.app.url_for('trackers.trackerfieldsjson',module=tracker.module,slug=tracker.slug),
                            'prePopulate':[ {'id':field.id,'name':field.name} for field in tracker.list_fields_list ]
                            },
                        'search_fields': {
                            'url': request.app.url_for('trackers.trackerfieldsjson',module=tracker.module,slug=tracker.slug),
                            'prePopulate':[ {'id':field.id,'name':field.name} for field in tracker.fields_from_list(tracker.search_fields) ]
                            },
                        'filter_fields': {
                            'url': request.app.url_for('trackers.trackerfieldsjson',module=tracker.module,slug=tracker.slug),
                            'prePopulate':[ {'id':field.id,'name':field.name} for field in tracker.fields_from_list(tracker.filter_fields) ]
                            },
                        'excel_fields': {
                            'url': request.app.url_for('trackers.trackerfieldsjson',module=tracker.module,slug=tracker.slug),
                            'prePopulate':[ {'id':field.id,'name':field.name} for field in tracker.fields_from_list(tracker.excel_fields) ]
                            },
                        'detail_fields': {
                            'url': request.app.url_for('trackers.trackerfieldsjson',module=tracker.module,slug=tracker.slug),
                            'prePopulate':[ {'id':field.id,'name':field.name} for field in tracker.fields_from_list(tracker.detail_fields) ]
                            },
                        }

    curuser = User.getuser(request['session']['user_id'])
    modules = curuser.rolemodules('Admin')
    return html(render(request,'generic/form.html',title=title,form=form,enctype='multipart/form-data',modules=modules,tokeninput=tokeninput),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})

@bp.route('/trackers')
@authorized(require_admin=True)
async def index(request):
    trackers = dbsession.query(Tracker)
    curuser = User.getuser(request['session']['user_id'])
    modules = []
    donefilter = False
    for m in dbsession.query(Tracker.module).distinct():
        if 'Admin' in curuser.moduleroles(m[0]):
            modules.append(m[0])
            if(request.args.get('module_filter') and request.args.get('module_filter')==m[0]):
                trackers = trackers.filter_by(module=m[0])
                donefilter = True
    if not donefilter:
        trackers = trackers.filter(Tracker.module.in_(modules))
    if request.args.get('q'):
        trackers = trackers.filter(or_(Tracker.title.ilike("%" + request.args.get('q') + "%"),Tracker.slug.ilike("%" + request.args.get('q') + "%")))
    paginator = Paginator(trackers, 10)
    return html(render(request,
        'generic/list.html',title='Trackers',editlink='trackers.viewtracker',addlink='trackers.create',maxlength=100,filter_fields=[{'field':'module','label':'Module','options':modules},],fields=[{'label':'Module','name':'module'},{'label':'Slug','name':'slug'},{'label':'Title','name':'title'},{'label':'List Fields','name':'list_fields'},{'label':'Published','name':'published'},{'label':'Require Login','name':'require_login'},{'label':'Allowed Roles','name':'allowed_roles'}],paginator=paginator,curpage=paginator.page(int(request.args['page'][0]) if 'page' in request.args else 1)),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})

@bp.route('/trackers/delete/<module>/<slug>',methods=['POST'])
@authorized(require_admin=True)
async def delete(request,module,slug=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    for update in tracker.dataupdates:
        if update.filename and os.path.exists(update.filename):
            os.remove(update.filename)
        if os.path.exists(os.path.join(uploadfolder,tracker.slug,'dataupdate',str(update.id))):
            os.rmdir(os.path.join(uploadfolder,tracker.slug,'dataupdate',str(update.id)))
        dbsession.delete(update)
        try:
            dbsession.commit()
        except Exception as inst:
            dbsession.rollback()
    for trackerstatus in tracker.statuses:
        print('deleting status' + str(trackerstatus))
        dbsession.delete(trackerstatus)
    for trackerrole in tracker.roles:
        print('deleting role' + str(trackerrole))
        dbsession.delete(trackerrole)
    for trackerfield in tracker.fields:
        print('deleting field' + str(trackerfield))
        dbsession.delete(trackerfield)
    for trackertransition in tracker.transitions:
        print('deleting transition' + str(trackertransition))
        dbsession.delete(trackertransition)
    try:
        dbsession.execute("drop table if exists " + tracker.data_table + ";")
        dbsession.execute("drop sequence if exists public." + tracker.data_table + "_id_seq;")
        dbsession.execute("drop table if exists " + tracker.update_table + ";")
        dbsession.execute("drop sequence if exists public." + tracker.update_table + "_id_seq;")
    except Exception as inst:
        dbsession.rollback()

    dbsession.delete(tracker)
    try:
        dbsession.commit()
    except Exception as inst:
        dbsession.rollback()
    return redirect(request.app.url_for('trackers.index'))

@bp.route('/system/cleardata/<module:string>/<slug:string>',methods=['POST'])
@authorized(object_type='dataupdate')
async def clearsystemdata(request,module,slug):
    returnurl = None
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    if 'returnurl' in request.form:
        returnurl = request.form.get('returnurl')
    if tracker:
        for update in tracker.dataupdates:
            if update.filename and os.path.exists(update.filename):
                os.remove(update.filename)
            if os.path.exists(os.path.join(uploadfolder,tracker.slug,'dataupdate',str(update.id))):
                os.rmdir(os.path.join(uploadfolder,tracker.slug,'dataupdate',str(update.id)))
            dbsession.delete(update)
        try:
            dbsession.commit()
        except Exception as inst:
            dbsession.rollback()
        try:
            dbsession.execute("delete from " + tracker.update_table)
            dbsession.execute("delete from " + tracker.data_table)
        except:
            dbsession.rollback()
        request['session']['flashmessage'] = 'Cleared data for ' + tracker.title
    if returnurl:
        return redirect(returnurl)
    else:
        return redirect(request.app.url_for('trackers.viewlist',module=module,slug=slug))


@bp.route('/system/<module:string>/<slug:string>/')
async def viewlist(request,module,slug=None):
    if slug==None:
        slug=module
        module='portal'
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    if 'accept' in request.headers and request.headers['accept']=='application/json':
        page = dbsession.query(Page).filter_by(module=module,slug=tracker.slug + '_json_list').first()
        if page:
            title = page.title
            ldict = locals()
            finalout = 'output=' + page.render(request,title=title,tracker=tracker).strip()
            exec(finalout,globals(),ldict)
            return jsonresponse(ldict['output'])
        else:
            title = tracker.title + '-List'
            ldict = locals()
            finalout = 'output=' + render(request,'trackers/viewlist.json',title=title,tracker=tracker).strip()
            exec(finalout,globals(),ldict)
            return jsonresponse(ldict['output'])
    else:
        page = dbsession.query(Page).filter_by(module=module,slug=tracker.slug + '_list').first()
        if page:
            title = page.title
            return html(page.render(request,title=title,tracker=tracker),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})
        else:
            title = tracker.title + '-List'
            return html(render(request,'trackers/viewlist.html',title=title,tracker=tracker),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})

@bp.route('/system/<module:string>/<slug:string>/excel')
async def listexcel(request,module,slug=None):
    if slug==None:
        slug=module
        module='portal'
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    curuser = None
    if 'user_id' in request['session']:
        curuser = dbsession.query(User).filter(User.id==request['session']['user_id']).first()
    records = tracker.records(curuser=curuser,request=request)
    wb = Workbook()
    ws = wb.active
    ws.title = tracker.title
    if records:
        xfields = []
        if tracker.excel_fields:
            xfields = tracker.fields_from_list(tracker.excel_fields)
        else:
            xfields = tracker.fields_from_list(tracker.list_fields)
        for i,f in enumerate(xfields):
            ws.cell(row=1,column=i+1,value=f.label)
        row=2
        for rec in records:
            for i,f in enumerate(xfields):
                ws.cell(row=row,column=i+1,value=str(f.disp_value(rec[f.name],request)))
            row+=1

    virtual_wb = save_virtual_workbook(wb)
    return raw(virtual_wb, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition':'inline;filename=' + slugify(tracker.title) + '.xlsx'})

@bp.route('/system/<module:string>/<slug:string>/method/<method:string>')
async def listmethod(request,module,slug,method):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    page = dbsession.query(Page).filter_by(module=module,slug=tracker.slug + '_' + method).first()
    if page:
        if page and page.runable:
            redirecturl=None
            results=None
            ldict = locals()
            try:
                exec(page.content,globals(),ldict)
            except:
                print("Got error running commands from page:" + str(page))
                print("Page content:" + str(page.content))
                print("Exception:" + str(sys.exc_info()[0]))
            if 'redirecturl' in ldict:
                redirecturl=ldict['redirecturl']
            if 'results' in ldict:
                results=ldict['results']
            if redirecturl:
                return redirect(redirecturl)
            elif results:
                return results
            else:
                return redirect('/')
        else:
            title = tracker.title + '-' + page.title
            return html(page.render(request,title=title,tracker=tracker),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})
    else:
        title = tracker.title + '- List'
        request['session']['flashmessage'] = 'Method ' + method + ' was not found for the tracker ' + tracker.title
        return html(render(request,'trackers/viewlist.html',title=title,tracker=tracker),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})

@bp.route('/system/<module:string>/<slug:string>/<id:int>',methods=['POST','GET'])
async def viewdetail(request,module,slug=None,id=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    curuser = None
    if 'user_id' in request['session']:
        curuser = dbsession.query(User).filter(User.id==request['session']['user_id']).first()
    record = None
    status = None
    page = None
    if request.method=='POST':
        return redirect(request.app.url_for('tracker.viewdetail',module=module,slug=slug))
    if(id):
        record = tracker.records(id,curuser=curuser,request=request)
        if record:
            status = tracker.status(record)
    if 'accept' in request.headers and request.headers['accept']=='application/json':
        if status:
            page = dbsession.query(Page).filter_by(module=module,slug=tracker.slug + '_view_' + status.name.lower().replace(' ','_') + '_json').first()
        if not page:
            page = dbsession.query(Page).filter_by(module=module,slug=tracker.slug + '_view_default_json').first()
        if page:
            title = page.title
            ldict = locals()
            finalout = 'output=' + page.render(request,record=record,title=title,tracker=tracker).strip()
            exec(finalout,globals(),ldict)
            return jsonresponse(ldict['output'])
        else:
            title = tracker.title + '-View'
            ldict = locals()
            finalout = 'output=' + render(request,'trackers/viewrecord.json',title=title,tracker=tracker,record=record).strip()
            exec(finalout,globals(),ldict)
            return jsonresponse(ldict['output'])
    else:
        if status:
            page = dbsession.query(Page).filter_by(module=module,slug=tracker.slug + '_view_' + status.name.lower().replace(' ','_')).first()
        if not page:
            page = dbsession.query(Page).filter_by(module=module,slug=tracker.slug + '_view_default').first()
        title = tracker.title + '-View'
        if page:
            return html(page.render(request,tracker=tracker,record=record,title=title),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})
        else:
            return html(render(request,'trackers/viewrecord.html',tracker=tracker,title=title,record=record),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})

@bp.route('/system/<module:string>/<slug:string>/method/<method:string>/<id:int>',methods=['POST','GET'])
async def detailmethod(request,module,slug,method,id):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    curuser = None
    if 'user_id' in request['session']:
        curuser = dbsession.query(User).filter(User.id==request['session']['user_id']).first()
    record = None
    status = None
    page = None
    if request.method=='POST':
        return redirect(request.app.url_for('tracker.detailmethod',module=module,slug=slug,method=method,id=id))
    if(id):
        record = tracker.records(id,curuser=curuser,request=request)
        if record:
            status = tracker.status(record)
    if status:
        page = dbsession.query(Page).filter_by(module=module,slug=tracker.slug + '_' + method + '_' + status.name.lower().replace(' ','_')).first()
        if not page:
            page = dbsession.query(Page).filter_by(module=module,slug=tracker.slug + '_view_' + status.name.lower().replace(' ','_')).first()
    if not page:
        page = dbsession.query(Page).filter_by(module=module,slug=tracker.slug + '_' + method + '_default').first()
        if not page:
            page = dbsession.query(Page).filter_by(module=module,slug=tracker.slug + '_view_default').first()
    if page:
        if page and page.runable:
            redirecturl=None
            results=None
            ldict = locals()
            try:
                exec(page.content,globals(),ldict)
            except:
                print("Got error running commands from page:" + str(page))
                print("Page content:" + str(page.content))
                print("Exception:" + str(sys.exc_info()[0]))
            if 'redirecturl' in ldict:
                redirecturl=ldict['redirecturl']
            if 'results' in ldict:
                results=ldict['results']
            if redirecturl:
                return redirect(redirecturl)
            elif results:
                return results
            else:
                return redirect('/')
        else:
            title = tracker.title + '- ' + page.title
            return html(page.render(request,tracker=tracker,record=record,title=title),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})
    else:
        request['session']['flashmessage'] = 'Method ' + method + ' was not found for the tracker ' + tracker.title
        title = tracker.title + '- View'
        return html(render(request,'trackers/viewrecord.html',tracker=tracker,title=title,record=record),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})

@bp.route('/system/<module:string>/<slug:string>/add',methods=['POST','GET'])
async def addrecord(request,module,slug=None):
    data = []
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    if tracker.default_new_transition:
        newtransition = dbsession.query(TrackerTransition).filter(TrackerTransition.tracker==tracker,TrackerTransition.name.ilike("%" + tracker.default_new_transition.strip() + "%")).first()
    else:
        newtransition = dbsession.query(TrackerTransition).filter(TrackerTransition.tracker==tracker,TrackerTransition.name.ilike("%new%")).first()
    if request.method=='POST':
        data = tracker.formsave(request.form,request)
        if newtransition.emails:
            newtransition.renderemails(request,data)
        if newtransition.postpage:
            ldict = locals()
            exec(newtransition.postpage,globals(),ldict)
            if 'output' in ldict and str(ldict['output']):
                return redirect(str(ldict['output']))
            else:
                return redirect(request.app.url_for('trackers.viewdetail',module=tracker.module,slug=tracker.slug,id=data['id']))
        else:
            if 'on_success' in request.form:
                return redirect(request.form['on_success'][0])
            else:
                return redirect(request.app.url_for('trackers.viewdetail',module=tracker.module,slug=tracker.slug,id=data['id']))
    page = dbsession.query(Page).filter_by(module=module,slug=tracker.slug + '_addrecord').first()
    title = tracker.title + "-Add Record"
    if newtransition:
        if page:
            return html(page.render(request,tracker=tracker,transition=newtransition,title=title),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})
        else:
            return html(render(request,'trackers/formrecord.html',tracker=tracker,transition=newtransition,title=title),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})
    else:
        request['session']['flashmessage'] = 'Default new transition not found'
        return redirect(request.app.url_for('pages.home'))

@bp.route('/system/<module:string>/<slug:string>/edit/<transition_id:int>/<record_id:int>',methods=['POST','GET'])
async def editrecord(request,module,slug=None,transition_id=None,record_id=None):
    tracker = dbsession.query(Tracker).filter_by(module=module,slug=slug).first()
    transition = None
    record = None
    curuser = None
    if 'user_id' in request['session']:
        curuser = dbsession.query(User).filter(User.id==request['session']['user_id']).first()
    if record_id:
        record = tracker.records(record_id,curuser=curuser,request=request)
    if transition_id:
        transition = dbsession.query(TrackerTransition).get(transition_id)
    if request.method=='POST':
        data = tracker.formsave(request.form,request,id=record_id)
        if transition.emails:
            transition.renderemails(request,data)
        if not data:
            return redirect(request.app.url_for('trackers.viewlist',module=tracker.module,slug=tracker.slug))
        if transition.postpage:
            ldict = locals()
            try:
                exec(transition.postpage,globals(),ldict)
            except Exception as inst:
                print("Error with postpage: " + str(inst))
            if 'output' in ldict and str(ldict['output']):
                return redirect(str(ldict['output']))
            else:
                return redirect(request.app.url_for('trackers.viewdetail',module=tracker.module,slug=tracker.slug,id=data['id']))
        else:
            return redirect(request.app.url_for('trackers.viewdetail',module=tracker.module,slug=tracker.slug,id=record_id))
    page = dbsession.query(Page).filter_by(module=module,slug=tracker.slug + '_editrecord').first()
    title = tracker.title + '-Edit Record'
    if page:
        return html(page.render(request,tracker=tracker,title=title,transition=transition,record=record),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})
    else:
        return html(render(request,'trackers/formrecord.html',tracker=tracker,title=title,transition=transition,record=record),headers={'X-Frame-Options':'deny','X-Content-Type-Options':'nosniff'})
