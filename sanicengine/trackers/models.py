# -*- coding: utf-8 -*-
"""Tracker models."""
import datetime

from sanicengine.database import ModelBase, dbsession, reference_col, executedb
from sqlalchemy import column, Column, ForeignKey, Integer, String, Text, Boolean, DateTime, Date, Table, MetaData, select, UniqueConstraint
from sqlalchemy.orm import relationship, backref
from sqlalchemy.sql import text
from sanicengine.template import render_string
from sanicengine.users.models import ModuleRole, User
from sanicengine.fileLinks.models import FileLink
from sanicengine import settings
from openpyxl import load_workbook
import json
import math
import os
import time
import traceback

class Tracker(ModelBase):
    """ 
    A class used to represent a Tracker

    Attributes
    ----------
    __tablename__ : str
        name of the table in the database
    id : int
        id of the tracker
    title : string
        the title of the tracker
    slug : string
        the slug of the tracker
    module : string
        the module the tracker belong to
    default_new_transition : string
        name of the transition to use to create a new record
    list_fields : string
        comma delimited list of fields (referred by name) to use when displaying 
        the records list
    search_fields : string
        comma delimited list of fields to use when making a search
    filter_fields : string
        comma delimited list of fields to use when making dropdown filter fields
    excel_fields : string
        comma delimited list of fields to use when generating an excel for download.
        If empty then will fallback on list from list_fields
    detail_fields : string
        comma delimited list of fields to use when displaying the details of the record.
        Will be overwritten by the status list of fields
    published : boolean
        indicate whether the tracker is published or not (default False)
    pagelimit : int
        default number of record to display per page in a list (default 10)
    require_login : boolean
        indicate whether the tracker requires the user to be logged in to be used
    allowed_roles : string
        comma delimited list of roles who can use the tracker.
        Anyone can use the tracker if not specified (anon users will rely on
        the require_login indicator though)
    publish_date : date
        date the tracker become published if published indicator is true.
        If null then tracker publishing just depend on published indicator
        and whether past expiry date if expire_date is specified
    expire_date : date
        date the tracker become expired if published indicator is true
        If null then never expire
    data_table_name : string
        name of the table the tracker data is kept in
    update_table_name : string
        name of the table the audit trail data is kept in
    list_order : string
        if exist will be appended onto " order by " portion of list query

    Methods
    -------
    load(slug,module=None)
        Static method to load the tracker based on slug and module if given
    newtransition()
        Property to get the transition to create a new record. Will search for transition
        named 'New' if default_new_transition is not specified
    is_published()
        Property to check whether the tracker should be published or not depending on the
        published, publish_date and expire_date values
    list_fields_list()
        Get an array of TrackerField from the valid field names specified in list_fields
    display_fields_list(record)
        Get an array of TrackerField to be displayed based on record status display_fields
        if specifed else fallback on the detail_fields from the tracker
    fields_from_list(field_list)
        Return array of TrackerField based on comma delimited string in field_list
    data_table()
        Property to get name of tracker data table. Will default to trak_{module}_{slug}_data
        if not specified
    update_table()
        Property to get name of audit data table. Will default to trak_{module}_{slug}_updates
        if not specified
    pages()
        Property to list all PortalPages that belong to the same module as current tracker
    field(name)
        Get TrackerField of this tracker with given name
    updatedb()
        Module to run to create/update the database tables based on values saved

    """

    __tablename__ = 'trackers'
    id = Column(Integer, primary_key=True)
    title = Column(String(200))
    slug = Column(String(100))
    module = Column(String(100),default='portal')
    default_new_transition = Column(String(100))
    list_fields = Column(Text)
    search_fields = Column(Text)
    filter_fields = Column(Text)
    excel_fields = Column(Text)
    detail_fields = Column(Text)
    published = Column(Boolean(),default=False)
    pagelimit = Column(Integer,default=10)
    require_login = Column(Boolean(),default=False)
    allowed_roles = Column(String(300))
    publish_date = Column(Date(),nullable=True)
    expire_date = Column(Date(),nullable=True)
    data_table_name = Column(String(200))
    update_table_name = Column(String(200))
    list_order = Column(String(200))

    __table_args__ = (
        UniqueConstraint(module, slug, name='tracker_module_slug_uidx'),
    )

    @property
    def data_table(self):
        """Name of table used by tracker to store records
        """

        if self.data_table_name:
            return self.data_table_name
        else:
            return "trak_" + self.module + "_" + self.slug + "_data"

    @property
    def update_table(self):
        """Name of table used by tracker to store the audit trail
        """

        if self.update_table_name:
            return self.update_table_name
        else:
            return "trak_" + self.module + "_" + self.slug + "_updates"

    @property
    def pages(self):
        """Array of pages with the same module as the tracker
        """

        from sanicengine.pages.models import Page
        return dbsession.query(Page).filter_by(module=self.module).all()

    @property
    def newtransition(self):
        """Transition to create a new record
        """

        if self.default_new_transition:
            return dbsession.query(TrackerTransition).filter_by(tracker=self, name=self.default_new_transition).first()
        else:
            return dbsession.query(TrackerTransition).filter_by(tracker=self, name='New').first()

    @property
    def is_published(self):
        """Determine whether tracker is published based on 
        published, publish_date and expire_date fields
        """

        if not self.published:
            return False
        else:
            toret = True
            curdate = datetime.date.today()
            if self.publish_date and self.publish_date>curdate:
                toret = False
            if self.expire_date and self.expire_date<curdate:
                toret = False
            return toret

    @property
    def list_fields_list(self):
        """Array of valid TrackerField from list_fields field
        """

        return self.fields_from_list(self.list_fields)

    def transition(self,trans_name,get_id=False):
        transition = dbsession.query(TrackerTransition).filter(TrackerTransition.name==trans_name,TrackerTransition.tracker==self).first()
        if transition:
            if get_id:
                return transition.id
            else:
                return transition
        else:
            return None

    def display_fields_list(self,record):
        """Array of valid TrackerField when displaying record based on 
        display_fields from the status or detail_fields from tracker
        """

        curstatus = self.status(record)
        if curstatus:
            if curstatus.display_fields:
                return self.fields_from_list(curstatus.display_fields)
            elif self.detail_fields:
                return self.fields_from_list(self.detail_fields)
        return []

    def __repr__(self):
        return self.title

    def load(slug,module=None):
        """Static method to load tracker based on slug and module

        If the argument `module` isn't passed in, trackers will
        be searched by slug only

        Parameters
        ----------
        slug : str
            The slug of the tracker to be loaded
        module : str, optional
            The module of the tracker to be loaded

        Returns
        -------
        Tracker
            The tracker found with the slug and module searched
        """

        tracker = dbsession.query(Tracker).filter_by(slug=slug)
        if module:
            tracker = tracker.filter_by(module=module)
        return tracker.first()

    def fields_from_list(self,field_list=None):
        """Method to get array of valid TrackerField from a comma delimited
        string of field names

        Parameters
        ----------
        field_list : str
            The comma delimited string of field names to split and find 
            valid TrackerField for the tracker

        Returns
        -------
        Array TrackerField
            Array of TrackerField found in the tracker from the string list
        """

        rfields = []
        if field_list:
            pfields = field_list.split(',')
            rfields = []
            for pfield in pfields:
                rfield = dbsession.query(TrackerField).filter_by(tracker=self,name=pfield.strip()).first()
                if rfield:
                    rfields.append(rfield)
        return rfields

    def field(self, name):
        """Find tracker field by it's name

        Parameters
        ----------
        name : str
            Name of the field to search for

        Returns
        -------
        TrackerField
            The TrackerField found with the name given
        """

        field = dbsession.query(TrackerField).filter_by(tracker=self,name=name.strip()).first()
        return field

    def updatedb(self):
        """Update the database to reflect the tracker

        Will create the data table if it doesn't exist yet. Will also create the audit trail table
        if it doesn't exist yet. Then will update the table for every field in the tracker
        """

        query = """
            do $$
            begin
            if (select not exists(select  1 from information_schema.sequences where sequence_schema = 'public' and sequence_name = '""" + self.data_table + """_id_seq'))
            then
                create sequence public.""" + self.data_table + """_id_seq
                increment 1
                minvalue 1
                maxvalue 9223372036854775807
                start 1
                cache 1;
            end if;
            if (select not exists(select  1 from information_schema.tables where table_schema = 'public' and table_name = '""" + self.data_table + """'))
            then
                create table public.""" + self.data_table + """(
                    id integer not null default nextval('""" + self.data_table + """_id_seq'::regclass),
                    record_status character varying(50) ,
                    batch_no integer
                );
            end if;
            if (select not exists(select  1 from information_schema.sequences where sequence_schema = 'public' and sequence_name = '""" + self.update_table + """_id_seq'))
            then
                create sequence public.""" + self.update_table + """_id_seq
                increment 1
                minvalue 1
                maxvalue 9223372036854775807
                start 1
                cache 1;
            end if;
            if (select not exists(select  1 from information_schema.tables where table_schema = 'public' and table_name = '""" + self.update_table + """'))
            then
                create table public.""" + self.update_table + """(
                    id integer not null default nextval('""" + self.update_table + """_id_seq'::regclass),
                    record_id integer,
                    user_id integer,
                    record_status character varying(50),
                    update_date timestamp,
                    description text
                );
            end if;
            end$$;
        """
        dbsession.execute(query)
        try:
            dbsession.commit()
        except Exception as inst:
            dbsession.rollback()
        for field in self.fields:
            field.updatedb()

    def first(self, query="1=1", qparams={}):
        """Will return the first record found from the tracker table

        Parameters
        ----------
        query : str
            Query string to be used after where in the sql statement
        qparams : dictionary
            Dictionary of values to be used in the query placeholders

        Returns
        -------
        dict
            Values of the first row found from the query
        """

        return executedb("select * from " + self.data_table + " where " + query,qparams).first()

    def query(self, query="1=1", qparams={}):
        """Will return the first record found from the tracker table

        Parameters
        ----------
        query : str
            Query string to be used after where in the sql statement
        qparams : dictionary
            Dictionary of values to be used in the query placeholders

        Returns
        -------
        dict
            Values of the first row found from the query
        """

        return executedb("select * from " + self.data_table + " where " + query,qparams)

    def saverecord(self, record, request=None):
        """Save the dictionary into the db

        Parameters
        ----------
        record : dict
            Dict of the values to be saved into the db
        request : request
            Request session to be used to determine identity of user

        Returns
        -------
        dict
            Values of the row once saved
        """
        
        data = None
        if 'id' in record:
            query = "update " + self.data_table + " set " + ",".join([ formfield + "=:" + formfield for formfield in record.keys()  ]) + " where id=:id returning *"
            qparams = { k:record[k] for k in record.keys() }
        else:
            query = "insert into " + self.data_table + "(" + ",".join(record.keys()) + ") values (" + ",".join([ ":" + formfield for formfield in record.keys()]) + ") returning *"
            qparams = { k:record[k] for k in record.keys() }
        try:
            data = dbsession.execute(query,qparams).fetchone()
            dbsession.commit()
        except Exception as inst:
            print("Error executing query:" + str(inst))
            dbsession.rollback()
        return data

    def formsave(self, form, request, id=None):
        """Save the information from submitted form into the db

        Parameters
        ----------
        form : dict
            Dict of values from the submitted form
        request : request
            Request session to be used to determine identity of user

        Returns
        -------
        dict
            Values of the row once saved
        """

        curuser = None
        transition = None
        if 'user_id' in request.ctx.session:
            curuser = dbsession.query(User).filter(User.id==request.ctx.session['user_id']).first()
        if 'transition_id' in form:
            transition = dbsession.query(TrackerTransition).get(form['transition_id'][0])
            if transition and transition.next_status:
                form['record_status'] = [transition.next_status.name,]
            del(form['transition_id'])
        oldrecord = None
        if id:
            oldrecord = self.records(id,curuser=curuser,request=request)
        elif 'id' in form:
            oldrecord = self.records(form['id'][0],curuser=curuser,request=request)
            del(form['id'])
        data = None
        if transition:
            if form['record_status'][0].lower()=='delete':
                if oldrecord:
                    request.ctx.session['flashmessage']="Deleted " + self.title + " " + str(oldrecord['id'])
                    try:
                        dbsession.execute("delete from " + self.update_table + " where record_id=:record_id",{'record_id':oldrecord['id']})
                    except Exception as inst:
                        dbsession.rollback()
                    try:
                        dbsession.execute("delete from " + self.data_table + " where id=:record_id",{'record_id':oldrecord['id']})
                    except Exception as inst:
                        dbsession.rollback()
                    return None
            else:
                fieldnames = []
                if 'record_status' in form:
                    fieldnames.append('record_status')
                for field in transition.edit_fields_list:
                    if field.default and field.name in form and form[field.name][0]=='systemdefault':
                        output = None
                        ldict = locals()
                        try:
                            exec(field.default,globals(),ldict)
                            output=ldict['output']
                            form[field.name][0]=output
                            fieldnames.append(field.name)
                        except Exception as inst:
                            from sanicengine.portalerrors.models import Error
                            Error.capture("Error exec default at " + request.url,str(field.default).replace("<","&lt;").replace(">","&gt;").replace("\n","<br>") + "<br><br>Error<br>==========<br>" + traceback.format_exc().replace("<","&lt;").replace(">","&gt;").replace("\n","<br>"))
                    elif field.field_type == 'boolean':
                        if field.name not in form:
                            form[field.name]=[False,]
                        else:
                            form[field.name]=[True,]
                        fieldnames.append(field.name)
                    elif field.field_type in ['file','picture','video']:
                        if request.files.get(field.name) and request.files.get(field.name).name:
                            filelink=FileLink()
                            dfile = request.files.get(field.name)
                            ext = dfile.type.split('/')[1]
                            if not os.path.exists(os.path.join('uploads',self.module)):
                                os.makedirs(os.path.join('uploads',self.module))
                            outfilename = str(int(time.time())) + dfile.name
                            dst = os.path.join('uploads',self.module,outfilename)
                            try:
                                # extract starting byte from Content-Range header string
                                range_str = request.headers['Content-Range']
                                start_bytes = int(range_str.split(' ')[1].split('-')[0])
                                with open(dst, 'ab') as f:
                                    f.seek(start_bytes)
                                    f.write(dfile.body)
                            except KeyError:
                                with open(dst, 'wb') as f:
                                    f.write(dfile.body)
                            filelink.filename = dfile.name
                            filelink.filepath = dst
                            filelink.module = self.module + "_data"
                            filelink.slug = outfilename.replace(' ','_').lower()
                            filelink.title = dfile.name
                            filelink.published = True
                            filelink.require_login = False
                            dbsession.add(filelink)
                            dbsession.commit()
                            form[field.name]=[filelink.id,]
                            fieldnames.append(field.name)
                    else:
                        if field.name not in form:
                            print(field.name + " not in form ")
                            form[field.name] = ["",]
                        fieldnames.append(field.name)
                if oldrecord:
                    query = "update " + self.data_table + " set " + ",".join([ formfield + "=:" + formfield for formfield in fieldnames  ]) + " where id=:record_id returning *"
                    ddata = { 'record_id':oldrecord['id'] }
                else:
                    query = "insert into " + self.data_table + " ( " + ",".join(fieldnames) + ") values (" + ",".join([ ":" + formfield for formfield in fieldnames  ]) + ") returning * "
                    ddata = {}
                try:
                    fdata = { field:form[field][0] for field in fieldnames }
                    ddata.update(fdata)
                    print("q:" + query)
                    data = dbsession.execute(query,ddata).fetchone()
                    dbsession.commit()
                except Exception as inst:
                    print("Error query:" + str(query) + " \n with data:" + str(ddata) + "\n" + str(inst))
                    dbsession.rollback()
                txtdesc = []
                if data:
                    if oldrecord:
                        for okey,oval in enumerate(oldrecord):
                            if oval!=data[okey]:
                                dfield = dbsession.query(TrackerField).filter_by(name=oldrecord.keys()[okey]).first()
                                if dfield:
                                    txtdesc.append('Updated ' + dfield.label + ' from ' + str(oval) + ' to ' + str(data[okey]))
                    desc = '<br>'.join(txtdesc)

                    query = " insert into " + self.update_table + " (record_id,user_id,record_status,update_date,description) values (:record_id,:user_id,:rec_status,now(),:desc) "
                    try:
                        dbsession.execute(query,{'record_id':data['id'],'user_id':curuser.id if curuser else None,'rec_status':data['record_status'],'desc':desc})
                        dbsession.commit()
                    except Exception as inst:
                        print("Error query:" + str(query))
                        dbsession.rollback()

        return data

    def pagelinks(self,curuser=None,request=None,cleared=False):
        # plo : page list offset
        # pll : page list limit
        pageamount = 0
        try:
            qtext, qparams = self.queryrules(curuser=curuser, request=request)
            recordamount = dbsession.execute("select count(*) as num from " + self.data_table + qtext,qparams).first()['num']
            pageamount = int(math.ceil(recordamount/(self.pagelimit if self.pagelimit else 10)))
        except Exception as inst:
            print("Exception getting record amount:" + str(inst))
            dbsession.rollback()
        links = []
        curoffset = 0
        if request:
            curoffset = int(request.args['plo'][0]) if 'plo' in request.args else 0
        curindex = int(curoffset/(self.pagelimit if self.pagelimit else 10))
        for x in range(0,pageamount):
            if x<11 or pageamount-x<11 or abs(x-curindex)<10:
                nextoffset=((x+1)*(self.pagelimit if self.pagelimit else 10) if x+1<pageamount else None)
                prevoffset=((x-1)*(self.pagelimit if self.pagelimit else 10) if x else None)
                offset=x*(self.pagelimit if self.pagelimit else 10)
                params = {'slug':self.slug,'pll':(self.pagelimit if self.pagelimit else 10)}
                if 'plq' in request.args:
                    params.update({'plq':request.args['plq'][0]})
                if self.filter_fields:
                    ffields = self.fields_from_list(self.filter_fields)
                    for ff in ffields:
                        pname = 'filter_' + ff.name
                        if pname in request.args:
                            params.update({pname:request.args[pname][0]})
                        elif pname + '_from' in request.args:
                            params.update({pname + '_from':request.args[pname+'_from'][0]})
                            if pname + '_to' in request.args:
                                params.update({pname + '_to':request.args[pname+'_to'][0]})
                params.update({'plo':offset})
                url=request.app.url_for('trackers.viewlist',module=self.module,**params)
                params.update({'plo':prevoffset})
                prevlink = (request.app.url_for('trackers.viewlist',module=self.module,**params)) if prevoffset else None
                params.update({'plo':nextoffset})
                nextlink = (request.app.url_for('trackers.viewlist',module=self.module,**params)) if nextoffset else None
                thislink = { 'url':url,'nextlink':nextlink,'prevlink':prevlink }
                links.append(thislink)
            else:
                thislink = { 'url':None,'nextlink':None,'prevlink':None }
                links.append(thislink)
        return curindex,links

    def queryrules(self,curuser=None,request=None,cleared=False):
        rules = ' where 1=1 '
        qparams = {}
        if request and 'plq' in request.args: # page list query exists
            if self.search_fields and len(request.args['plq'][0]):
                sfields = self.fields_from_list(self.search_fields)
                sqs = []
                for sfield in sfields:
                    qrule,qparam = sfield.queryvalue(request.args['plq'][0])
                    sqs.append(qrule)
                    qparams.update(qparam)
                if len(sqs):
                    rules += ' and (' + ' or '.join(sqs) + ')'

        if self.filter_fields:
            ffields = self.fields_from_list(self.filter_fields)
            fqs = []
            for ffield in ffields:
                if 'filter_' + ffield.name in request.args and request.args['filter_'+ffield.name][0]!='plall':
                    qrule,qparam = ffield.queryvalue(request.args['filter_'+ffield.name][0],equals=True)
                    fqs.append(qrule)
                    qparams.update(qparam)
                elif 'filter_' + ffield.name + '_from' in request.args:
                    qrule,qparam = ffield.queryvalue([request.args['filter_' + ffield.name + '_from'][0],request.args['filter_' + ffield.name + '_to'][0]])
                    fqs.append(qrule)
                    qparams.update(qparam)
            if len(fqs):
                rules += ' and (' + ' and '.join(fqs) + ')'

        if not cleared:
            rrules = self.rolesrule(curuser,request)
            if rrules:
                rules += ' and (' + rrules + ') '
        return rules, qparams

    def records(self,id=None,curuser=None,request=None,cleared=False,offset=None,limit=None):
        results = []
        qparams = {}
        if id:
            if cleared:
                sqltext = text("select * from " + self.data_table +  " where id=:id")
            else:
                sqltext = text("select * from " + self.data_table +  " where id=:id and (" + self.rolesrule(curuser,request) + ")")
            sqltext = sqltext.bindparams(id=id)
        else:
            limitstr = ''
            offsetstr = ''
            if limit:
                limitstr = ' limit ' + str(limit)
            if offset:
                offsetstr = ' offset ' + str(offset)
            qtext, qparams = self.queryrules(curuser=curuser,request=request,cleared=cleared)
            lorder = ''
            if self.list_order:
                lorder = ' order by ' + self.list_order + ' '

            strsql = "select * from " + self.data_table + qtext + lorder + limitstr + offsetstr
            sqltext = text(strsql)
        try:
            results = dbsession.execute(sqltext,qparams)
        except Exception as inst:
            dbsession.rollback()
        if id:
            drows = None
            for row in results:
                drows = row
            results = drows
        return results

    def rolesrule(self,curuser,request):
        """Return the query rules based on roles of the curuser

        Parameters
        ----------
        curuser : User
            user that needs to be checked against
        request : request
            Request session to be used to determine identity of user

        Returns
        -------
        string
            Value of the query to be used to check the role
        """

        if self.userroles(curuser,request=request):
            rolesrule = '1=1'
        else:
            queryroles = dbsession.query(TrackerRole).filter_by(tracker=self,role_type='query').all()
            rolesrule = '1=0'
            rulestext = [ render_string(request,role.compare) for role in queryroles ]
            if len(rulestext):
                trolesrule = ' or '.join(rulestext)
                if len(trolesrule):
                    rolesrule = trolesrule
        return rolesrule

    def status(self,record):
        if record and 'record_status' in record and record['record_status']:
            status = dbsession.query(TrackerStatus).filter_by(tracker=self,name=record['record_status']).first()
            if status:
                return status
        return None

    def userroles(self,curuser,record=None,request=None):
        """Return an array of user roles based on the tracker and record
        if given

        Parameters
        ----------
        curuser : User
            user that needs to be checked against
        record : dict
            data from the tracker table
        request : request
            Request session to be used to determine identity of user

        Returns
        -------
        array
            Array of roles the user has on the tracker and record if exists
        """
        croles = []
        for role in self.roles:
            if role.role_type=='module':
                usermoduleroles = dbsession.query(ModuleRole).filter(ModuleRole.module==self.module,ModuleRole.user==curuser,ModuleRole.role==role.name).all()
                if len(usermoduleroles):
                    croles.append(role)
            elif record:
                rolesrule = render_string(request,role.compare)
                sqltext = "select id from " + self.data_table + " where id=:record_id and " + rolesrule
                try:
                    results = dbsession.execute(sqltext,{"record_id":record['id']})
                    if len(results):
                        croles.append(role)
                except Exception as inst:
                    dbsession.rollback()
        return croles

    def activetransitions(self,record,curuser,request):
        status = self.status(record)
        roles = self.userroles(curuser,record,request=request)
        atransitions = []
        if status:
            transitions = dbsession.query(TrackerTransition).filter(TrackerTransition.prev_status==status,TrackerTransition.tracker==self).all()
            if len(transitions) and len(roles):
                for t in transitions:
                    for r in roles:
                        if r in t.roles:
                            atransitions.append(t)
        return atransitions

    def recordupdates(self,record,curuser):
        updates = []
        if record:
            try:
                updates = dbsession.execute("select * from " + self.update_table + " where record_id=:record_id order by update_date desc",{"record_id":record['id']})
            except Exception as inst:
                print("Error updating record:" + str(inst))
                dbsession.rollback()
        return updates

    def recordvalue(self,record,field,request=None):
        ftags = field.split('.')
        if len(ftags)==1:
            dfield = dbsession.query(TrackerField).filter(TrackerField.tracker==self,TrackerField.name==ftags[0]).first()
            if dfield:
                return dfield.disp_value(record[ftags[0]],request=request)

class TrackerField(ModelBase):
    __tablename__ = 'tracker_fields'
    id = Column(Integer, primary_key=True)
    name = Column(String(50))
    label = Column(String(50))
    field_type = Column(String(20))
    widget = Column(String(20))
    obj_table = Column(String(50))
    obj_field = Column(String(100))
    obj_filter = Column(String(200))
    session_override = Column(String(100))
    get_override = Column(String(100))
    default = Column(Text())
    options = Column(Text())
    foreignfields = []

    tracker_id = reference_col('trackers')
    tracker = relationship('Tracker',backref='fields')

    def __repr__(self):
        return self.label

    def load_foreign_fields(self):
        if self.field_type=='hasMany' and self.obj_table and self.obj_field:
            self.foreignfields = []
            tt = self.obj_table.split('.')
            dtracker = dbsession.query(Tracker).filter_by(module=tt[0],slug=tt[1]).first()
            if dtracker:
                for f in self.obj_field.split(','):
                    ddm = dbsession.query(TrackerField).filter_by(tracker=dtracker,name=f.strip()).first()
                    if ddm:
                        self.foreignfields.append(ddm)


    def listnames(datas,sep=','):
        toret = []
        for d in datas.split(sep):
            try:
                transition = dbsession.query(TrackerField).get(int(d))
                if transition:
                    toret.append(transition.name)
            except Exception as e:
                print("Cannot get field with id " + str(d) + " :" + str(e))
        return toret

    def obj_fields(self):
        if self.obj_field:
            return self.obj_field.split(',')

    def main_obj_field(self):
        if self.obj_field:
            return self.obj_fields()[0]

    def filter_options(self,curuser=None,request=None,cleared=False):
        values = []
        try:
            qtext, qparams = self.tracker.queryrules(curuser=curuser,request=request,cleared=cleared)
            if self.field_type == 'object' and self.obj_table and self.obj_field:
                cobj_field = self.main_obj_field()
                values = dbsession.execute("select distinct cur." + self.name + " as val, ref." + cobj_field + " as label from " + self.tracker.data_table + " cur, " + self.obj_table + " ref " + qtext + " and ref.id=cur." + self.name + " order by ref." + cobj_field,qparams)
            elif self.field_type == 'user':
                values = dbsession.execute("select distinct cur." + self.name + " as val, ref.name as label from " + self.tracker.data_table + " cur, users ref " + qtext + " and ref.id=cur." + self.name + " order by ref." + self.name,qparams)
            else:
                values = dbsession.execute("select distinct " + self.name + " as val from " + self.tracker.data_table + " " + qtext + " order by " + self.name,qparams)
        except Exception as inst:
            print("Error getting filter:" + str(inst))
            dbsession.rollback()
        return values

    def disp_options(self):
        if self.options:
            output=None
            try:
                ldict = locals()
                exec(field.options,globals(),ldict)
                output=ldict['output']
            except:
                output=[ k.strip() for k in self.options.split(',') ]
            return output
        else:
            return None

    def disp_value(self, value, request=None):
        if value:
            if self.field_type=='object':
                sqlq = "select " + self.main_obj_field() + " from " + self.obj_table + " where id=:obj_id"
                try:
                    result = dbsession.execute(sqlq,{'obj_id':value})
                    for r in result:
                        return r[0]
                except Exception as inst:
                    print("Error running query:" + str(inst))
                    dbsession.rollback()
            elif self.field_type=='treenode':
                sqlq = "select string_agg(cnode.title,'->' order by cnode.lft) from tree_nodes cnode,(select id,lft,rgt,tree_id from tree_nodes where id=:node_id) nleaf where cnode.lft<=nleaf.lft and cnode.rgt>=nleaf.rgt and cnode.tree_id=nleaf.tree_id group by nleaf.lft,nleaf.rgt,nleaf.id,nleaf.tree_id"
                try:
                    result = dbsession.execute(sqlq,{'node_id':value})
                    for r in result:
                        return r[0]
                except Exception as inst:
                    print("Error running query:" + str(inst))
                    dbsession.rollback()
            elif self.field_type=='user':
                sqlq = "select name from users where id=:user_id"
                try:
                    result = dbsession.execute(sqlq,{'user_id':value})
                    for r in result:
                        return r[0]
                except Exception as inst:
                    print("Error running query:" + str(inst))
                    dbsession.rollback()
            elif self.field_type == 'picture' or self.widget== 'picture':
                filelink = dbsession.query(FileLink).get(value)
                if filelink:
                    url=request.app.url_for('fileLinks.download',module=filelink.module,slug=filelink.slug)
                    value = "<img src='" + url + "'>"
            elif self.field_type == 'video' or self.widget== 'video':
                filelink = dbsession.query(FileLink).get(value)
                if filelink:
                    url=request.app.url_for('fileLinks.download',module=filelink.module,slug=filelink.slug)
                    value = "<video controls><source src='" + url + "'></video>"
            elif self.field_type in ['location','map']:
                mlat = value.split(',')
                try:
                    if len(mlat)>1 and settings.GOOGLE_MAPS_API:
                        mapstr = "<div id='" + self.name + "_div' style='height: 400px;'></div>"
                        mapstr += "<script>var " + self.name + "_map," + self.name + "_marker; "
                        mapstr += "function init_" + self.name + "_map(){ " + self.name + "_map = new google.maps.Map(document.getElementById('" + self.name + "_div'), { center: {lat: " + mlat[0] + ",lng:" + mlat[1] + "}, zoom: 13 }); " + self.name + "_marker = new google.maps.Marker({ position: {lat:" + mlat[0] + ",lng:" + mlat[1] + " }, map: " + self.name + "_map }); }</script>"
                        return mapstr
                    else:
                        return 'Map failed to render for location: ' + value
                except:
                    return 'Map failed to render for location: ' + value
            
        if self.field_type=='boolean':
            result = ['False','True']
            if self.disp_options() and len(self.disp_options())>0:
                result = self.disp_options()
            if not value:
                value = result[0]
            else:
                value = result[1]
        return value

    def queryvalue(self, value, equals=False):
        """Will return name of query to run with placeholders for variables
        e.g: if field name is staff_id, will return "staff_id=:staff_id" and
        dictionary with value mapped to staff_id 

        Parameters
        ----------
        value : str or list
            value is the value used to compare which would be put into qparams to send as parameter
                if equals is true then it would be put in between % before putting into parameter
                if the field type is a date then value is a list consisting of from and to date to compare
        equals : bool
            by default certain string based field would compare with like %% but equals true would force it to be compared as equal

        Returns
        -------
        toret
            Query string to be run
        qparams
            Parameterized values to be run in query
        """

        toret = self.name + "=:" + self.name
        qparams = { self.name: value }
        if self.field_type in ['string','text','location','map'] and not equals:
            qparams = { self.name: "%" + str(value) + "%" }
            toret = self.name + " ilike :" + self.name
        if self.field_type in ['date','datetime']:
            qparams = { self.name + '_from': value[0], self.name + '_to': value[1] }
            toret = self.name + ">=:" + self.name + "_from and " + self.name + "<=:" + self.name +"_to"
        return toret, qparams

    def sqlvalue(self, value):
        if value is not None:
            if self.field_type in ['string','text','date','datetime','location','map']:
                return "'" + str(value).replace("'","''") + "'"
            elif self.field_type in ['integer','number','object','treenode','user','file','picture','video']:
                return str(value)
            elif self.field_type=='boolean':
                if value:
                    return 'true'
                else:
                    return 'false'
        else:
            return "null"
        return ''

    def dbcolumn(self):
        return column(self.name)

    def db_field_type(self):
        if self.field_type=='string':
            return 'character varying(200)'
        elif self.field_type in ['location','map']:
            return 'character varying(200)'
        elif self.field_type=='text':
            return 'text'
        elif self.field_type=='integer':
            return 'integer'
        elif self.field_type=='object':
            return 'integer'
        elif self.field_type=='treenode':
            return 'integer'
        elif self.field_type=='user':
            return 'integer'
        elif self.field_type=='belongsTo':
            return 'integer'
        elif self.field_type=='file':
            return 'integer'
        elif self.field_type=='picture':
            return 'integer'
        elif self.field_type=='video':
            return 'integer'
        elif self.field_type=='number':
            return 'double precision'
        elif self.field_type=='date':
            return 'date'
        elif self.field_type=='datetime':
            return 'timestamp'
        elif self.field_type=='boolean':
            return 'boolean'

    def updatedb(self):
        if not self.field_type=='hasMany':
            query = """
                do $$
                begin 
                    IF not EXISTS (SELECT column_name 
                        FROM information_schema.columns 
                        WHERE table_schema='public' and table_name='""" + self.tracker.data_table + """' and column_name='""" + self.name + """') THEN
                            alter table public.""" + self.tracker.data_table + ' add column "' + self.name + '" ' + self.db_field_type() + """ null ;
                    end if;
                end$$;
            """
            try:
                dbsession.execute(query)
                dbsession.commit()
            except Exception as inst:
                print("Error running updating db:" + str(inst))
                dbsession.rollback()

class TrackerRole(ModelBase):
    __tablename__ = 'tracker_roles'
    id = Column(Integer, primary_key=True)
    name = Column(String(50))
    role_type = Column(String(10))
    compare = Column(Text,nullable=True)

    tracker_id = reference_col('trackers')
    tracker = relationship('Tracker',backref='roles')

    def __repr__(self):
        return self.name

class TrackerDataUpdate(ModelBase):
    __tablename__ = 'tracker_data_update'
    id = Column(Integer, primary_key=True)
    status = Column(String(50))
    created_date = Column(DateTime)
    filename = Column(String(200))
    data_params = Column(Text,nullable=True)

    tracker_id = reference_col('trackers')
    tracker = relationship('Tracker',backref='dataupdates')

    def __repr__(self):
        return self.tracker.title + ' ' + str(self.created_date)

    async def run(self):
        wb = load_workbook(filename = self.filename)
        ws = wb.active
        datas = json.loads(self.data_params)
        fields = []
        optype = 'insert'
        searchfield = []
        for field in self.tracker.fields:
            if field.name != 'id' and datas[field.name + '_column'][0]!='ignore':
                fields.append(field)
            if field.name + '_update' in datas:
                optype = 'update'
                searchfield.append(field)
        rows = []
        headerend = 1
        if optype == 'insert':
            recstatusq = ''
            newtransition = None
            if 'record_status' not in fields:
                recstatusq = ',record_status'
                newtransition = dbsession.query(TrackerTransition).filter(TrackerTransition.tracker==self.tracker,TrackerTransition.name==self.tracker.default_new_transition).first()
            query = 'insert into ' + self.tracker.data_table + ' (' + ','.join([ '"' + f.name + '"' for f in fields ]) + ',batch_no' + recstatusq + ') values (' + ','.join([ ':' + f.name for f in fields ]) + ',:batch_no' + ( ',:record_status' if newtransition else '') + ')'
            qparams = []
            for i in range(headerend+1,ws.max_row+1):
                if i>headerend:
                    cellrows = {f.name:(ws.cell(column=int(datas[f.name + '_column'][0]),row=i).value if datas[f.name + '_column'][0]!='custom' else datas[f.name + '_custom'][0]) for f in fields }
                    cellrows.update( { 'batch_no':self.id } )
                    if 'record_status' not in fields:
                        if newtransition and newtransition.next_status:
                            cellrows.update( {'record_status':newtransition.next_status.name })
                    qparams.append(cellrows)
            try:
                executedb(query,qparams)
                dbsession.commit()
                self.status = 'Uploaded ' + str(len(qparams)) + ' rows'
                dbsession.add(self)
                dbsession.commit()
            except Exception as inst:
                print("Error inserting data:" + str(inst))
                dbsession.rollback()
        else:
            allresults = []
            for i in range(headerend+1,ws.max_row+1):
                if i>headerend:
                    fieldupdates = []
                    wquery = {}
                    for f in searchfield:
                        cellrow = ws.cell(column=int(datas[f.name + '_column'][0]),row=i)
                        fieldupdates.append( f.name + '=:' + f.name )
                        wquery.update( { f.name: cellrow.value } )
                    queryrow = ' where ' + ' and '.join(fieldupdates)

                    sqltext = text("select id from " + self.tracker.data_table + queryrow)
                    try:
                        result = executedb(sqltext,wquery)
                        gotdata = False
                        for r in result:
                            gotdata = True
                    except Exception as inst:
                        print("Error running query:" + str(inst))
                        dbsession.rollback()
                    
                    if gotdata:
                        query = 'update ' + self.tracker.data_table + ' set ' 
                        fieldinfos = []
                        uquery = {}
                        for f in fields:
                            if not f in searchfield:
                                fieldinfos.append( f.name + '=:' + f.name )
                                if datas[f.name + '_column'][0]!='custom':
                                    cellrow = ws.cell(column=int(datas[f.name + '_column'][0]),row=i)
                                    uquery.update({ f.name:cellrow.value })
                                else:
                                    cellrow = datas[f.name + '_custom'][0]
                                    uquery.update({ f.name:cellrow })
                        uquery.update(wquery)
                        query += ','.join(fieldinfos) + queryrow + ' returning *'
                        try:
                            result = executedb(query,uquery)
                        except Exception as inst:
                            print("Error runing query:" + str(inst))
                            dbsession.rollback()
                    else:
                        celldatas = {}
                        for f in fields:
                            if datas[f.name + '_column'][0]!='custom':
                                celldatas[f.name] = ws.cell(column=int(datas[f.name + '_column'][0]),row=i).value
                            else:
                                celldatas[f.name] = datas[f.name + '_custom'][0]
                        celldatas['batch_no'] = self.id
                        query = 'insert into ' + self.tracker.data_table + ' (' + ','.join([ f.name for f in fields ]) + ',batch_no) values (' + ','.join([ ':' + f.name for f in fields ]) + ',:batch_no)'
                        query += ' returning *'
                        result = executedb(query,celldatas)
                    try:
                        dbsession.commit()
                        allresults.append('Uploaded ' + str(result.first()))
                    except Exception as inst:
                        dbsession.rollback()
            try:
                self.status = 'Updated ' + str(len(allresults)) + ' rows'
                dbsession.add(self)
                dbsession.commit()
            except Exception as inst:
                dbsession.rollback()
        return True

class TrackerStatus(ModelBase):
    __tablename__ = 'tracker_statuses'
    id = Column(Integer, primary_key=True)
    name = Column(String(50))
    display_fields = Column(Text)

    tracker_id = reference_col('trackers')
    tracker = relationship('Tracker',backref='statuses')

    def __repr__(self):
        return self.name

transition_roles = Table('transition_roles',ModelBase.metadata,
        Column('transition_id',Integer,ForeignKey('tracker_transitions.id')),
        Column('role_id',Integer,ForeignKey('tracker_roles.id'))
        )

transition_emails = Table('transition_emails',ModelBase.metadata,
        Column('transition_id',Integer,ForeignKey('tracker_transitions.id')),
        Column('emailtemplate_id',Integer,ForeignKey('emailtemplates.id'))
        )

class TrackerTransition(ModelBase):
    __tablename__ = 'tracker_transitions'
    id = Column(Integer, primary_key=True)
    name = Column(String(50))
    label = Column(String(50))
    email_ids = Column(Text)
    display_fields = Column(Text)
    edit_fields = Column(Text)
    postpage = Column(Text)

    roles = relationship('TrackerRole',secondary=transition_roles,backref=backref('transitions',lazy='dynamic'))

    emails = relationship('EmailTemplate',secondary=transition_emails,backref=backref('transitions',lazy='dynamic'))

    tracker_id = reference_col('trackers')
    tracker = relationship('Tracker',backref='transitions')

    prev_status_id = reference_col('tracker_statuses',nullable=True)
    prev_status = relationship('TrackerStatus',backref='prev_transitions',foreign_keys=[prev_status_id])

    next_status_id = reference_col('tracker_statuses',nullable=True)
    next_status = relationship('TrackerStatus',backref='next_transitions',foreign_keys=[next_status_id])

    def __repr__(self):
        return self.name

    @property
    def edit_fields_list(self):
        pfields = []
        if self.edit_fields:
            pfields = self.edit_fields.split(',') if self.edit_fields else []
        rfields = []
        for pfield in pfields:
            rfield = dbsession.query(TrackerField).filter_by(tracker=self.tracker,name=pfield.strip()).first()
            if rfield:
                rfields.append(rfield)
        return rfields

    def display_fields_list(self):
        pfields = []
        if self.display_fields:
            pfields = self.display_fields.split(',')
        rfields = []
        for pfield in pfields:
            rfield = dbsession.query(TrackerField).filter_by(tracker=self.tracker,name=pfield.strip()).first()
            if rfield:
                rfields.append(rfield)
        return rfields

    def renderemails(self,request,data):
        if self.emails:
            for email in self.emails:
                email.renderemail(request,data=data)
